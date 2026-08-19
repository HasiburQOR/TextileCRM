"""
Commercial Invoice / Packing List exports (PDF + Excel).

Layout follows the SUMAIYA-style combined document this business actually
issues: company letterhead (name, ID/registration no, address, contact
person, TEL), the invoice meta (DATE / Inv No), one combined COMMERCIAL
INVOICE / PACKING LIST table with twenty-three columns per product line —
including an embedded product photo per row, and carton SIZE as a grouped
header spanning its L/W/H sub-columns — a totals block that shows its own
arithmetic (bill + commission = total; each bank transfer summed =
received), and the bank details customers wire against.

Every label is bilingual (English/Georgian, BR-33/FR-60-63): the renderers
take `lang="en"|"ka"` and the export view maps `?lang=` onto it. Georgian
needs a font with Mkhedruli glyphs — reportlab's built-in Helvetica has
none, nor even № — so DejaVu Sans ships in backend/assets/fonts and the
PDF falls back to English labels when it can't be loaded; Excel has no
such limit because fonts render client-side.

Nothing here computes a commercial figure. Line-level arithmetic (total
qty, amount, CBM, per-line weights) is settled in
`apps.invoicing.services.compute_line_item` at generation time and frozen
on the row; document totals come from `Invoice.document_totals()`. This
module only lays those numbers out — so the PDF, the Excel and the API can
never quote three different answers.
"""

import io
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
# `openpyxl.utils` re-exports only the cell helpers — pixels_to_EMU lives in
# the units submodule and importing it from the package raises ImportError,
# which takes this whole module (and with it every invoice export, PDF
# included) down at import time.
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as RLImage
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.core.models import CompanyProfile
from apps.invoicing.models import Invoice
from apps.sourcing.models import ImageLabel

HEADER_FILL = "1E2937"
LINE_HEADER_FILL = "E5E7EB"
TOTALS_FILL = "F1F5F9"
LINE_HEADER_FILL_HEX = "#" + LINE_HEADER_FILL

# Photo thumbnails are generated at this pixel width. Full-resolution
# product photos are routinely 2-4 MB each; a 190-line invoice embedding
# the originals would be a ~400 MB download that no mail server accepts.
PHOTO_PX = 220
PHOTO_MM = 14  # printed size in the PDF's Foto column (18mm cell − 2×2mm padding)
# The letterhead is printed ~180mm wide, so it needs far more pixels than a
# 16mm thumbnail before it starts looking soft.
LOGO_PX = 1400

# Column set for the line table, one entry per column of the reference
# document: (column key, label key into _T). Kept in one place because the
# PDF, the Excel sheet and their header rows must not drift apart.
COLUMNS = [
    ("no", "no"),  # auto row number; a shipping mark rides underneath
    ("photo", "photo"),
    ("description", "description"),
    ("brand", "brand"),
    ("details", "details"),
    ("ctn", "ctn"),
    ("qtyPerCtn", "qtyPerCtn"),
    ("totalQty", "totalQty"),
    # Unit Price prints per line: the price captured at Sourcing Intake
    # (ProductVariant.unitPrice), carried onto the Packing List's carton row
    # (PackingCarton.unitPrice) and locked onto this invoice line at
    # generation, so the document quotes the same figure the sourcing intake
    # and packing list do.
    ("unitPrice", "unitPrice"),
    # Per-line AMOUNT (unitPrice × totalQty) so the buyer can verify the
    # arithmetic that feeds the TOTAL VALUE at the bottom.
    ("amount", "amount"),
    ("nwPerCtn", "nwPerCtn"),
    ("totalNw", "totalNw"),
    ("gwPerCtn", "gwPerCtn"),
    ("totalGw", "totalGw"),
    # Carton SIZE prints as a grouped header spanning three sub-columns
    # (L / W / H, centimetres), matching the reference's merged-cell layout.
    ("sizeL", "sizeL"),
    ("sizeW", "sizeW"),
    ("sizeH", "sizeH"),
    ("cbmPerCtn", "cbmPerCtn"),
    ("cbm", "cbm"),
    ("material", "material"),
    # The reference's rightmost lettered column (A2, K13…) maps to OUR real
    # identifiers — style_no / pattern_no, e.g. "MRF25 / MR12528" — never a
    # fabricated code scheme (Reference_Numbers_Identifier_System.md).
    ("style", "style"),
]
KEYS = [key for key, _ in COLUMNS]
SIZE_KEYS = ("sizeL", "sizeW", "sizeH")
SIZE_GROUP_START = KEYS.index(SIZE_KEYS[0])  # first of the three SIZE columns


# Bilingual labels (BR-33 / FR-60-63): English + Georgian, one dict both
# renderers and every block (header, meta, totals, bank) share, so a
# translation can never drift between the PDF and the Excel.
_T = {
    # Document headings
    "doc_title_1": {"en": "COMMERCIAL INVOICE", "ka": "კომერციული ინვოისი"},
    "doc_title_2": {"en": "PACKING LIST", "ka": "შეფუთვის სია"},
    "date": {"en": "DATE", "ka": "თარიღი"},
    "invoice_no": {"en": "Inv No", "ka": "ინვოისი №"},
    "consignee": {"en": "To / Consignee:", "ka": "მიმღები / Consignee:"},
    "po_line": {"en": "PO / Sister Profile", "ka": "PO / ორდერი"},
    "id_reg_no": {"en": "ID / Registration No", "ka": "ID / რეგისტრაციის №"},
    "contact": {"en": "Contact", "ka": "კონტაქტი"},
    "tel": {"en": "TEL", "ka": "ტელ."},
    # Line-table columns
    "no": {"en": "№", "ka": "№"},
    "photo": {"en": "Foto", "ka": "ფოტო"},
    "description": {"en": "Description", "ka": "აღწერა"},
    "brand": {"en": "Brand", "ka": "ბრენდი"},
    "details": {"en": "Details", "ka": "დეტალები"},
    "ctn": {"en": "CTN", "ka": "კონტ."},
    "qtyPerCtn": {"en": "QTY/CTN", "ka": "რაოდ./კონტ."},
    "totalQty": {"en": "T. QTY", "ka": "სულ რაოდ."},
    "unitPrice": {"en": "Unit Price", "ka": "ერთეულის ფასი"},
    "amount": {"en": "AMOUNT", "ka": "თანხა"},
    "nwPerCtn": {"en": "N.W", "ka": "წმ. წონა"},
    "totalNw": {"en": "T/N.W", "ka": "სულ წმინდა"},
    "gwPerCtn": {"en": "G.W", "ka": "ბრუტო წონა"},
    "totalGw": {"en": "T/G.W", "ka": "სულ ბრუტო"},
    "size_group": {"en": "SIZE (cm)", "ka": "ზომა (სმ)"},
    "sizeL": {"en": "L", "ka": "L"},
    "sizeW": {"en": "W", "ka": "W"},
    "sizeH": {"en": "H", "ka": "H"},
    "cbmPerCtn": {"en": "CBM/ctn", "ka": "მ³/კონტ."},
    "cbm": {"en": "CBM", "ka": "მ³"},
    "material": {"en": "Material", "ka": "მასალა"},
    "style": {"en": "Style No", "ka": "სტილის №"},
    "total": {"en": "TOTAL", "ka": "სულ"},
    # Totals block
    "total_value": {"en": "TOTAL VALUE", "ka": "სულ ღირებულება"},
    "total_value_converted": {"en": "TOTAL VALUE (CONVERTED)", "ka": "სულ ღირებულება (კონვერტირებული)"},
    "commission": {"en": "COMMISSION", "ka": "კომისია"},
    "total_bill_commission": {"en": "TOTAL BILL + COMMISSION", "ka": "სულ ანგარიში + კომისია"},
    "bank_transfer": {"en": "BANK TRANSFER", "ka": "საბანკო გადარიცხვა"},
    "exchange_rate": {"en": "EXCHANGE RATE", "ka": "ვალუტის კურსი"},
    "locked_at_generation": {"en": "locked at generation", "ka": "ფიქსირებულია გენერაციის მომენტში"},
    "total_qty": {"en": "TOTAL QTY", "ka": "სულ რაოდენობა"},
    "total_ctns": {"en": "TOTAL CTNS/BAG", "ka": "სულ კონტ./ტომარა"},
    "total_cube": {"en": "TOTAL CUBE", "ka": "სულ მოცულობა"},
    "total_net_weight": {"en": "TOTAL NET WEIGHT", "ka": "სულ წმინდა წონა"},
    "total_gross_weight": {"en": "TOTAL GROSS WEIGHT", "ka": "სულ ბრუტო წონა"},
    # Bank block
    "bank_details": {"en": "BANK DETAILS", "ka": "საბანკო დეტალები"},
    "bank_name": {"en": "Bank Name", "ka": "ბანკის სახელი"},
    "account_title": {"en": "Account Title", "ka": "ანგარიშის მფლობელი"},
    "account_no": {"en": "Account No", "ka": "ანგარიშის №"},
    "swift_code": {"en": "SWIFT Code", "ka": "SWIFT კოდი"},
    "bank_address": {"en": "Bank Address", "ka": "ბანკის მისამართი"},
    "not_configured": {
        "en": "Not yet configured — set them in Company Profile.",
        "ka": "არ არის მითითებული — შეავსეთ კომპანიის პროფილი.",
    },
    "authorized_signatory": {"en": "Authorized Signatory", "ka": "უფლებამოსილი ხელმომწერი"},
}


def t(key: str, lang: str = "en") -> str:
    return _T[key].get(lang) or _T[key]["en"]


# ── Fonts ───────────────────────────────────────────────────────────
# Helvetica (reportlab's built-in) has neither Georgian Mkhedruli nor even
# the № glyph. DejaVu Sans ships with the app (backend/assets/fonts) and
# covers Latin + Georgian, so the bilingual PDF can be a real document
# rather than a page of black boxes.
FONT_DIR = Path(__file__).resolve().parents[2] / "assets" / "fonts"
_font_cache = None  # ("DejaVu", "DejaVu-Bold") once loaded, else False


def _register_fonts():
    """Idempotent, memoized. Returns (regular, bold) family names, or None
    when the TTFs are missing/corrupt — callers then stay on Helvetica and
    the PDF falls back to English labels."""
    global _font_cache
    if _font_cache is None:
        regular, bold = FONT_DIR / "DejaVuSans.ttf", FONT_DIR / "DejaVuSans-Bold.ttf"
        try:
            pdfmetrics.registerFont(TTFont("DejaVu", str(regular)))
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", str(bold)))
            # So `<b>…</b>` inside a Paragraph resolves to the bold face.
            pdfmetrics.registerFontFamily(
                "DejaVu", normal="DejaVu", bold="DejaVu-Bold", italic="DejaVu", boldItalic="DejaVu-Bold"
            )
            _font_cache = ("DejaVu", "DejaVu-Bold")
        except Exception:
            _font_cache = False
    return _font_cache if _font_cache else None


def _money(value) -> str:
    return f"{float(value or 0):,.2f}"


def _num(value, places=2) -> str:
    if value in (None, ""):
        return "—"
    return f"{float(value):,.{places}f}"


def _esc(value) -> str:
    """reportlab parses Paragraph text as mini-HTML — an unescaped `&` in a
    brand name ("Jack & Jones", which is all over a real invoice) aborts the
    render."""
    return escape(str(value if value is not None else ""))


def invoice_filename(invoice: Invoice, ext: str) -> str:
    return f"{invoice.invoiceNo}.{ext}"


# ── Photos ─────────────────────────────────────────────────────────────

def _thumbnail(image_field, max_px: int = PHOTO_PX) -> io.BytesIO | None:
    """A small RGB JPEG of a product photo, or None if there isn't one /
    it can't be read.

    Never raises: a missing or corrupt image file must degrade to an empty
    Foto cell, not take down the whole invoice download. Converts to RGB
    because a PNG with alpha cannot be saved as JPEG, and CMYK JPEGs (which
    phone-to-desktop workflows produce) render with inverted colours.
    """
    if not image_field:
        return None
    try:
        with image_field.open("rb"):
            img = PILImage.open(image_field)
            img.load()
    except Exception:
        return None
    try:
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        img.thumbnail((max_px, max_px), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=82)
        buf.seek(0)
        return buf
    except Exception:
        return None


def _line_photo(line) -> io.BytesIO | None:
    """The photo for one invoice line, taken from the source Product's own
    image gallery — upload a photo against the product and it appears here,
    with nothing to attach per invoice.

    Prefers the image labelled "Product Overall", falling back to the
    earliest upload. ProductImage is ordered by createdAt, so a bare
    `.first()` would pick whatever happened to be uploaded first — often a
    fabric close-up or a back label, which tells a buyer nothing about what
    they're being invoiced for.

    `product` is SET_NULL, so a line whose product was later deleted simply
    has no photo; every financial column is copied onto the line itself and
    stays correct regardless.
    """
    product = line.product
    if not product:
        return None
    images = list(product.images.all())
    if not images:
        return None
    chosen = next((img for img in images if img.label == ImageLabel.PRODUCT_OVERALL), images[0])
    return _thumbnail(chosen.image)


# ── Shared content blocks ──────────────────────────────────────────────

def _company_lines(company: CompanyProfile, lang: str = "en") -> list[str]:
    """Letterhead text, in the reference header's order: name, ID /
    registration no, address, contact person, TEL. Also the fallback when
    no logo has been uploaded — an invoice with no identity on it at all
    would be unusable."""
    lines = []
    if company.name:
        lines.append(company.name)
    if company.tagline:
        lines.append(company.tagline)
    if company.registrationNo:
        lines.append(f"{t('id_reg_no', lang)}: {company.registrationNo}")
    if company.addressLine:
        lines.append(company.addressLine.replace("\n", ", "))
    if company.contactPerson:
        lines.append(f"{t('contact', lang)}: {company.contactPerson}")
    if company.phone:
        lines.append(f"{t('tel', lang)}: {company.phone}")
    if company.email:
        lines.append(f"E-mail: {company.email}")
    return lines


def _consignee_lines(invoice: Invoice, lang: str = "en") -> list[str]:
    """Build consignee/buyer block for the printed document.

    When per-invoice ``InvoiceBuyerDetails`` are present (typed at creation
    time), those lines are used instead of the SisterProfile → BuyerProfile
    hierarchy — the operator explicitly fills them in for the document.
    """
    buyer_details = getattr(invoice, "_buyer_details_cache", None)
    if buyer_details is None:
        try:
            buyer_details = invoice.buyerDetails
            # Cache it for repeated access (e.g. PDF + XLSX on same instance).
            invoice._buyer_details_cache = buyer_details
        except Exception:
            buyer_details = None
    if buyer_details:
        lines: list[str] = []
        if buyer_details.companyName:
            lines.append(buyer_details.companyName)
        if buyer_details.idNumber:
            lines.append(f"I/D NO: {buyer_details.idNumber}")
        if buyer_details.address:
            lines.append(f"ADD: {buyer_details.address}")
        if buyer_details.cityCountry:
            lines.append(buyer_details.cityCountry)
        if buyer_details.contactPerson:
            lines.append(buyer_details.contactPerson)
        if buyer_details.phone:
            lines.append(f"TEL: {buyer_details.phone}")
        # Custom label/value pairs
        for field in (buyer_details.customFields or []):
            label = field.get("label", "")
            value = field.get("value", "")
            if label and value:
                lines.append(f"{label}: {value}")
        if lines:
            return lines

    sister_profile = invoice.sisterProfile
    buyer = sister_profile.buyerProfile
    lines = [buyer.name]
    if buyer.branding:
        lines.append(buyer.branding)
    if buyer.contactInfo:
        lines.extend(part for part in buyer.contactInfo.splitlines() if part.strip())
    lines.append(f"{t('po_line', lang)}: {sister_profile.poReference or sister_profile.referenceCode}")
    return lines


def _bank_rows(company: CompanyProfile, lang: str = "en") -> list[tuple[str, str]]:
    rows = [
        (t("bank_name", lang), company.bankName),
        (t("account_title", lang), company.bankAccountTitle),
        (t("account_no", lang), company.bankAccountNo),
        (t("swift_code", lang), company.bankSwiftCode),
        (t("bank_address", lang), (company.bankAddress or "").replace("\n", ", ")),
    ]
    return [(label, value) for label, value in rows if value]


def _totals_rows(invoice: Invoice, lang: str = "en") -> list[tuple[str, str]]:
    """The summary block.

    The money side now shows TWO totals: TOTAL VALUE in the invoice's
    source (line-item) currency, and TOTAL VALUE (CONVERTED) in the
    target currency — so the buyer sees both figures regardless of what
    currency the lines are priced in. The exchange rate that connects
    them is also printed, so the conversion is auditable on paper.
    """
    totals = invoice.document_totals()
    source = invoice.sourceCurrency or ""
    target = invoice.targetCurrency or ""
    paid = invoice.total_paid()

    rows = [(t("total_value", lang), f"{_money(invoice.grand_total())} {source}".strip())]

    # Show the converted total if we have a target currency and a rate.
    converted = invoice.convertedTotal
    if target and converted:
        rows.append((t("total_value_converted", lang), f"{_money(converted)} {target}"))

    # Show the exchange rate so the conversion is auditable on paper.
    if invoice.exchangeRateValueLocked:
        rows.append((t("exchange_rate", lang), invoice.rate_label()))

    if paid:
        payments = list(invoice.payments.all().order_by("paymentDate", "createdAt"))
        parts = " + ".join(
            _money(p.amount) if (not source or p.currency == source) else f"{_money(p.amount)} {p.currency}"
            for p in payments
        )
        rows.append((t("bank_transfer", lang), f"{parts} = {_money(paid)} {source}".strip()))

    rows += [
        (t("total_qty", lang), f"{totals['totalQty']:,} PCS"),
        (t("total_ctns", lang), f"{totals['totalCtn']:,}"),
        (t("total_cube", lang), f"{_num(totals['totalCbm'], 3)} CBM"),
        (t("total_net_weight", lang), f"{_num(totals['totalNetWeight'])} KG"),
        (t("total_gross_weight", lang), f"{_num(totals['totalGrossWeight'])} KG"),
    ]
    return rows


def _style_no(line) -> str:
    """The Style No cell: our real style/pattern identifiers, e.g.
    "MRF25 / MR12528"."""
    parts = [p for p in (line.styleItemCode, line.patternNo) if p]
    return " / ".join(parts) or "—"


def _cell_values(line, index: int) -> dict:
    """One invoice line as {column key: display value}. `photo` and `no`
    are handled separately by each renderer since the two embed images /
    compose the row number + shipping mark differently."""
    return {
        "description": line.description or "—",
        "brand": line.brand or "—",
        "details": line.colorSizeNote or "—",
        "ctn": f"{line.ctn:,}",
        "qtyPerCtn": f"{line.qtyPerCtn:,}",
        "totalQty": f"{line.totalQty:,}",
        "unitPrice": _num(line.unitPrice),
        "amount": _num(line.amount),
        "nwPerCtn": _num(line.netWeightPerCtn),
        "totalNw": _num(line.netWeight),
        "gwPerCtn": _num(line.grossWeightPerCtn),
        "totalGw": _num(line.grossWeight),
        "sizeL": _num(line.ctnLengthCm, 1) if line.ctnLengthCm else "—",
        "sizeW": _num(line.ctnWidthCm, 1) if line.ctnWidthCm else "—",
        "sizeH": _num(line.ctnHeightCm, 1) if line.ctnHeightCm else "—",
        "cbmPerCtn": _num(line.cbmPerCtn, 4),
        "cbm": _num(line.cbm, 4),
        "material": line.material or "—",
        "style": _style_no(line),
    }


# ── PDF ────────────────────────────────────────────────────────────────
# Landscape A4: twenty-three columns of invoice + packing detail do not fit
# portrait at a readable type size.

PDF_COLUMN_WIDTHS = {
    "no": 8, "photo": 18, "description": 22, "brand": 15, "details": 22,
    "ctn": 8, "qtyPerCtn": 9, "totalQty": 9, "unitPrice": 14, "amount": 16,
    "nwPerCtn": 8, "totalNw": 10, "gwPerCtn": 8, "totalGw": 10,
    "sizeL": 9, "sizeW": 9, "sizeH": 9,
    "cbmPerCtn": 10, "cbm": 10, "material": 21, "style": 21,
}  # sums to 266mm — the printable width of landscape A4 at 14mm margins.
# The 16mm the Amount column takes is reclaimed from the descriptive
# columns around it, so the table still spans the full printable width.


def render_invoice_pdf(invoice: Invoice, lang: str = "en") -> bytes:
    fonts = _register_fonts()
    # Georgian text is impossible without a Mkhedruli-capable font; rather
    # than print a document full of black boxes, fall back to English.
    lang = lang if (lang == "ka" and fonts) else "en"
    regular, bold = fonts if fonts else ("Helvetica", "Helvetica-Bold")

    company = CompanyProfile.load()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"Commercial Invoice {invoice.invoiceNo}",
    )
    styles = getSampleStyleSheet()
    company_style = ParagraphStyle("Company", parent=styles["Title"], fontName=bold, fontSize=17, spaceAfter=1, leading=20)
    company_sub = ParagraphStyle("CompanySub", parent=styles["Normal"], fontName=regular, fontSize=8.5, alignment=1, leading=11)
    title_style = ParagraphStyle("DocTitle", parent=styles["Title"], fontName=bold, fontSize=12.5, alignment=1, spaceBefore=6, spaceAfter=1, leading=15)
    block_style = ParagraphStyle("Block", parent=styles["Normal"], fontName=regular, fontSize=8.5, leading=11)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontName=regular, fontSize=6.5, leading=8, wordWrap="CJK")
    header_style = ParagraphStyle("CellHeader", parent=cell_style, fontName=bold, alignment=1)

    story = []

    # ── Letterhead: the uploaded logo, or typeset text if none ──────
    logo = _thumbnail(company.logo, LOGO_PX) if company.logo else None
    if logo:
        # Scale to the full printable width, preserving aspect ratio — a
        # letterhead banner is much wider than tall.
        reader = PILImage.open(logo)
        width_px, height_px = reader.size
        logo.seek(0)
        # Top-centre placement, aspect ratio preserved in BOTH directions:
        # cap the height at 26mm and shrink the width to match when the cap
        # bites — clamping only the height squashed any logo taller than a
        # ~7:1 banner into a distorted 180×26mm strip.
        ratio = (height_px / width_px) if width_px else 0
        if not ratio:  # unreadable dimensions — degrade, don't fail the export
            ratio, draw_width = 0.11, 180 * mm
        else:
            draw_width = min(180 * mm, (26 * mm) / ratio)
        letterhead = RLImage(logo, width=draw_width, height=draw_width * ratio)
        # reportlab's Image flowable carries no hAlign attribute, so without
        # this it renders flush LEFT — the letterhead belongs top-centre.
        letterhead.hAlign = "CENTER"
        story.append(letterhead)
        story.append(Spacer(1, 4))
    else:
        # An uploaded letterhead IS the header — these text lines are the
        # fallback for when there isn't one, not a caption underneath it.
        # (Printing both duplicates the address on every real letterhead,
        # which already carries it.)
        for i, text in enumerate(_company_lines(company, lang)):
            story.append(Paragraph(_esc(text), company_style if i == 0 else company_sub))

    story.append(Spacer(1, 6))

    # ── Consignee (left) and invoice meta (right) ───────────────────
    # The reference header shows DATE, then Inv No — status is a workflow
    # concern, not something the buyer's customs broker needs on paper.
    consignee = "<br/>".join(_esc(line) for line in _consignee_lines(invoice, lang))
    meta = "<br/>".join([
        f"<b>{t('date', lang)}:</b> {invoice.createdAt.strftime('%d.%m.%Y')}",
        f"<b>{t('invoice_no', lang)}:</b> {_esc(invoice.invoiceNo)}",
    ])
    head_table = Table(
        [[Paragraph(f"<b>{_esc(t('consignee', lang))}</b><br/>{consignee}", block_style), Paragraph(meta, block_style)]],
        colWidths=[180 * mm, 86 * mm],
    )
    head_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(head_table)
    # The reference stacks the two document titles, centered.
    story.append(Paragraph(_esc(t("doc_title_1", lang)), title_style))
    story.append(Paragraph(_esc(t("doc_title_2", lang)), title_style))

    # ── Line table ──────────────────────────────────────────────────
    # Two header rows so SIZE can span its L/W/H sub-columns; every other
    # header spans both rows vertically, as in the reference's merged cells.
    no_label = t("no", lang) if fonts else "No."  # Helvetica has no №
    row_one, row_two = [], []
    for key, label_key in COLUMNS:
        if key in SIZE_KEYS:
            row_one.append(Paragraph(_esc(t("size_group", lang)) if key == "sizeL" else "", header_style))
            row_two.append(Paragraph(_esc(t(label_key, lang)), header_style))
        else:
            row_one.append(Paragraph(_esc(no_label if key == "no" else t(label_key, lang)), header_style))
            row_two.append("")
    data = [row_one, row_two]
    for index, line in enumerate(invoice.lineItems.all(), start=1):
        values = _cell_values(line, index)
        row = []
        for key, _ in COLUMNS:
            if key == "photo":
                photo = _line_photo(line)
                row.append(RLImage(photo, width=PHOTO_MM * mm, height=PHOTO_MM * mm) if photo else "")
            elif key == "no":
                # Auto row number, with the shipping mark painted on the
                # carton riding underneath — both identify the row on the
                # dock, and neither is worth a column of its own.
                cell = str(index)
                if line.markRef:
                    cell += f"<br/><font size=5.5>{_esc(line.markRef)}</font>"
                row.append(Paragraph(cell, cell_style))
            else:
                row.append(Paragraph(_esc(values[key]), cell_style))
        data.append(row)

    table = Table(
        data,
        colWidths=[PDF_COLUMN_WIDTHS[key] * mm for key, _ in COLUMNS],
        repeatRows=2,
    )
    span_commands = []
    for position, (key, _) in enumerate(COLUMNS):
        if key == "sizeL":
            span_commands.append(("SPAN", (position, 0), (position + len(SIZE_KEYS) - 1, 0)))
        elif key not in SIZE_KEYS:
            span_commands.append(("SPAN", (position, 0), (position, 1)))
    numeric_from, numeric_to = KEYS.index("ctn"), KEYS.index("cbm")
    table.setStyle(TableStyle(span_commands + [
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor(LINE_HEADER_FILL_HEX)),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C7CDD5")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 2), (1, -1), "CENTER"),
        ("ALIGN", (numeric_from, 2), (numeric_to, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(table)
    story.append(Spacer(1, 8))

    # ── Totals + bank details, side by side ─────────────────────────
    totals_data = [[Paragraph(f"<b>{_esc(label)}</b>", block_style), Paragraph(_esc(value), block_style)]
                   for label, value in _totals_rows(invoice, lang)]
    totals_table = Table(totals_data, colWidths=[52 * mm, 86 * mm])
    totals_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))

    bank_rows = _bank_rows(company, lang)
    if bank_rows:
        bank_data = [[Paragraph(f"<b>{_esc(t('bank_details', lang))}</b>", block_style), ""]]
        bank_data += [[Paragraph(f"<b>{_esc(label)}</b>", block_style), Paragraph(_esc(value), block_style)]
                      for label, value in bank_rows]
    else:
        bank_data = [[Paragraph(
            f"<b>{_esc(t('bank_details', lang))}</b><br/>{_esc(t('not_configured', lang))}", block_style), ""]]
    bank_table = Table(bank_data, colWidths=[34 * mm, 90 * mm])
    bank_table.setStyle(TableStyle([
        ("SPAN", (0, 0), (-1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(LINE_HEADER_FILL_HEX)),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))

    footer = Table([[totals_table, bank_table]], colWidths=[140 * mm, 126 * mm])
    footer.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0)]))

    # ── Seal + signature ─────────────────────────────────────────────
    # Bottom-right, above the "Authorized Signatory" caption — the
    # conventional spot on a commercial invoice, and distinct from the
    # letterhead (top) so the two never compete for the same banner.
    # Grouped into one KeepTogether with the totals/bank footer above: on a
    # tight page break, the signature belongs with the figures it's signing
    # off on, not orphaned alone at the top of the next page.
    footer_block = [footer]
    seal = _thumbnail(company.sealSignature, LOGO_PX) if company.sealSignature else None
    if seal:
        reader = PILImage.open(seal)
        width_px, height_px = reader.size
        seal.seek(0)
        ratio = (height_px / width_px) if width_px else 0.4
        draw_width = min(60 * mm, (26 * mm) / ratio) if ratio else 60 * mm
        signature_image = RLImage(seal, width=draw_width, height=draw_width * ratio)
        signature_image.hAlign = "RIGHT"
        footer_block.append(Spacer(1, 10))
        footer_block.append(signature_image)
    footer_block.append(Spacer(1, 2))
    footer_block.append(Paragraph(_esc(t("authorized_signatory", lang)), ParagraphStyle(
        "Signatory", parent=block_style, alignment=2,  # right
    )))
    story.append(KeepTogether(footer_block))

    doc.build(story)
    return buf.getvalue()


# ── Excel ──────────────────────────────────────────────────────────────

XLSX_COLUMN_WIDTHS = {
    "no": 7, "photo": 14, "description": 38, "brand": 20, "details": 30,
    "ctn": 7, "qtyPerCtn": 9, "totalQty": 10, "unitPrice": 10, "amount": 14,
    "nwPerCtn": 9, "totalNw": 10, "gwPerCtn": 9, "totalGw": 10,
    "sizeL": 7, "sizeW": 7, "sizeH": 7,
    "cbmPerCtn": 10, "cbm": 10, "material": 28, "style": 20,
}
PHOTO_ROW_HEIGHT = 78  # points — tall enough for a ~1in thumbnail


def _logo_center_anchor(image_px: int) -> tuple[int, int]:
    """Column index + EMU offset that horizontally centres an `image_px`
    wide picture across the sheet's fixed 21-column grid. openpyxl anchors
    pictures to cells and "A1" means flush LEFT, so centring takes a
    OneCellAnchor with an offset into the grid. Excel's own chars→pixels
    approximation (round(width×7)+5) is close enough that the residual
    error is invisible on a letterhead."""
    widths_px = [round(XLSX_COLUMN_WIDTHS[key] * 7) + 5 for key, _ in COLUMNS]
    left = max(0, (sum(widths_px) - image_px) // 2)
    col = 0
    while col < len(widths_px) - 1 and left >= widths_px[col]:
        left -= widths_px[col]
        col += 1
    return col, pixels_to_EMU(left)


def render_invoice_xlsx(invoice: Invoice, lang: str = "en") -> bytes:
    company = CompanyProfile.load()
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    line_header_font = Font(bold=True, size=9)
    line_header_fill = PatternFill("solid", fgColor=LINE_HEADER_FILL)
    totals_fill = PatternFill("solid", fgColor=TOTALS_FILL)
    thin = Side(style="thin", color="C7CDD5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_col = len(COLUMNS)
    last_col_letter = get_column_letter(last_col)

    row = 1
    # openpyxl anchors images to cells and keeps the stream open until
    # save(), so every buffer has to outlive this function's locals.
    open_streams: list[io.BytesIO] = []

    # ── Letterhead: the uploaded banner, or typeset text if none ────
    logo_placed = False
    if company.logo:
        logo = _thumbnail(company.logo, LOGO_PX)
        if logo:
            open_streams.append(logo)
            picture = XLImage(logo)
            ratio = picture.height / picture.width if picture.width else 0.25
            picture.width = 420
            picture.height = int(420 * ratio)
            # Top-centre, same as the PDF — "A1" would sit the banner flush
            # left of the grid it belongs over.
            col, col_off = _logo_center_anchor(picture.width)
            picture.anchor = OneCellAnchor(
                _from=AnchorMarker(col=col, colOff=col_off, row=0, rowOff=0),
                ext=XDRPositiveSize2D(pixels_to_EMU(picture.width), pixels_to_EMU(picture.height)),
            )
            ws.add_image(picture)
            ws.row_dimensions[1].height = max(40, picture.height * 0.78)
            row, logo_placed = 2, True
    if not logo_placed:
        for i, text in enumerate(_company_lines(company, lang)):
            ws.merge_cells(f"A{row}:{last_col_letter}{row}")
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = title_font if i == 0 else bold if i == 1 else Font(size=9)
            cell.alignment = Alignment(horizontal="center")
            row += 1
    row += 1

    # ── Consignee + meta ────────────────────────────────────────────
    # DATE then Inv No, as on the reference header — no status on paper.
    consignee = _consignee_lines(invoice, lang)
    meta = [
        (t("date", lang), invoice.createdAt.strftime("%d.%m.%Y")),
        (t("invoice_no", lang), invoice.invoiceNo),
    ]
    ws.cell(row=row, column=1, value=t("consignee", lang)).font = bold
    for offset, text in enumerate(consignee):
        ws.cell(row=row + offset, column=2, value=text)
    meta_col = last_col - 3  # top-right, where the reference puts it
    for offset, (label, value) in enumerate(meta):
        ws.cell(row=row + offset, column=meta_col, value=label).font = bold
        ws.cell(row=row + offset, column=meta_col + 1, value=value)
    row += max(len(consignee), len(meta)) + 1

    # The reference stacks the two document titles, centered.
    for title_text in (t("doc_title_1", lang), t("doc_title_2", lang)):
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        title = ws.cell(row=row, column=1, value=title_text)
        title.font = header_font
        title.fill = header_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1
    row += 1

    # ── Line table ──────────────────────────────────────────────────
    # Two header rows so SIZE can span its L/W/H sub-columns (merged
    # horizontally); every other header spans both rows vertically.
    header_row = row
    size_start_col = KEYS.index(SIZE_KEYS[0]) + 1
    size_end_col = KEYS.index(SIZE_KEYS[-1]) + 1
    for col, (key, label_key) in enumerate(COLUMNS, start=1):
        if col < size_start_col or col > size_end_col:
            ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row + 1, end_column=col)
            ws.cell(row=header_row, column=col, value=t(label_key, lang))
    ws.merge_cells(start_row=header_row, start_column=size_start_col, end_row=header_row, end_column=size_end_col)
    ws.cell(row=header_row, column=size_start_col, value=t("size_group", lang))
    for key in SIZE_KEYS:
        ws.cell(row=header_row + 1, column=KEYS.index(key) + 1, value=t(key, lang))
    for r in (header_row, header_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = line_header_font
            cell.fill = line_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 2

    keys = KEYS
    numeric_keys = {"ctn", "qtyPerCtn", "totalQty", "unitPrice", "amount", "nwPerCtn",
                    "totalNw", "gwPerCtn", "totalGw", "sizeL", "sizeW", "sizeH",
                    "cbmPerCtn", "cbm"}
    numeric_source = {
        "ctn": lambda li: li.ctn, "qtyPerCtn": lambda li: li.qtyPerCtn, "totalQty": lambda li: li.totalQty,
        "unitPrice": lambda li: li.unitPrice, "amount": lambda li: li.amount,
        "nwPerCtn": lambda li: li.netWeightPerCtn, "totalNw": lambda li: li.netWeight,
        "gwPerCtn": lambda li: li.grossWeightPerCtn, "totalGw": lambda li: li.grossWeight,
        "sizeL": lambda li: li.ctnLengthCm, "sizeW": lambda li: li.ctnWidthCm, "sizeH": lambda li: li.ctnHeightCm,
        "cbmPerCtn": lambda li: li.cbmPerCtn, "cbm": lambda li: li.cbm,
    }

    for index, line in enumerate(invoice.lineItems.all(), start=1):
        values = _cell_values(line, index)
        for col, key in enumerate(keys, start=1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=key in ("description", "details", "material", "style"))
            if key == "photo":
                photo = _line_photo(line)
                if photo:
                    open_streams.append(photo)
                    picture = XLImage(photo)
                    # Square-ish box; keeps rows uniform regardless of the
                    # photo's own aspect ratio.
                    picture.width = picture.height = 92
                    ws.add_image(picture, f"{get_column_letter(col)}{row}")
                continue
            if key == "no":
                # Row number with the shipping mark beneath it — both
                # identify the row on the dock.
                cell.value = f"{index}\n{line.markRef}" if line.markRef else index
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                continue
            # Numbers go in as numbers, not preformatted strings — an
            # invoice you can't re-total in Excel isn't much of an export.
            if key in numeric_keys:
                raw = numeric_source[key](line)
                cell.value = float(raw or 0)
                if key in SIZE_KEYS:
                    cell.number_format = "0.0"  # carton dimensions: 60.0, never 6,00.0
                else:
                    cell.number_format = "#,##0.000000" if key == "cbmPerCtn" else (
                        "#,##0.0000" if key == "cbm" else "#,##0.00" if key not in ("ctn", "qtyPerCtn", "totalQty") else "#,##0"
                    )
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = values[key]
        ws.row_dimensions[row].height = PHOTO_ROW_HEIGHT
        row += 1

    # Totals row directly under the table, as SUM formulas so the file
    # stays live if someone edits a quantity in their own copy.
    first_line_row, last_line_row = header_row + 2, row - 1
    if last_line_row >= first_line_row:
        ws.cell(row=row, column=1, value=t("total", lang)).font = bold
        for col, key in enumerate(keys, start=1):
            if key in ("ctn", "totalQty", "amount", "totalNw", "totalGw", "cbm"):
                letter = get_column_letter(col)
                cell = ws.cell(row=row, column=col, value=f"=SUM({letter}{first_line_row}:{letter}{last_line_row})")
                cell.font = bold
                cell.fill = totals_fill
                cell.number_format = "#,##0.0000" if key == "cbm" else "#,##0.00" if key in ("totalNw", "totalGw") else "#,##0"
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = border
    row += 2

    # ── Totals block + bank details ─────────────────────────────────
    for label, value in _totals_rows(invoice, lang):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value=t("bank_details", lang)).font = bold
    ws.cell(row=row, column=1).fill = line_header_fill
    row += 1
    bank_rows = _bank_rows(company, lang)
    if bank_rows:
        for label, value in bank_rows:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=value)
            row += 1
    else:
        ws.cell(row=row, column=1, value=t("not_configured", lang))
    row += 2

    # ── Seal + signature ─────────────────────────────────────────────
    # Bottom-right, mirroring the PDF: image, then a caption row beneath it.
    if company.sealSignature:
        seal = _thumbnail(company.sealSignature, LOGO_PX)
        if seal:
            open_streams.append(seal)
            picture = XLImage(seal)
            ratio = picture.height / picture.width if picture.width else 0.4
            picture.width = 170
            picture.height = int(170 * ratio)
            seal_col = last_col - 2  # bottom-right, clear of the bank/totals block
            col_off_px = max(0, (XLSX_COLUMN_WIDTHS[KEYS[seal_col - 1]] * 7 + 5 - picture.width) // 2)
            picture.anchor = OneCellAnchor(
                _from=AnchorMarker(col=seal_col - 1, colOff=pixels_to_EMU(col_off_px), row=row - 1, rowOff=0),
                ext=XDRPositiveSize2D(pixels_to_EMU(picture.width), pixels_to_EMU(picture.height)),
            )
            ws.add_image(picture)
            ws.row_dimensions[row].height = max(40, picture.height * 0.78)
            row += 1
    caption = ws.cell(row=row, column=last_col - 2, value=t("authorized_signatory", lang))
    caption.alignment = Alignment(horizontal="center")
    caption.font = bold

    for col, (key, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = XLSX_COLUMN_WIDTHS[key]
    ws.freeze_panes = ws.cell(row=header_row + 2, column=1)
    ws.print_title_rows = f"{header_row}:{header_row + 1}"

    out = io.BytesIO()
    wb.save(out)
    for stream in open_streams:
        stream.close()
    return out.getvalue()
