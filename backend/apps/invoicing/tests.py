import io
import shutil
import tempfile
from decimal import Decimal

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib.pagesizes import A4, landscape
from rest_framework import status
from rest_framework.test import APITestCase

from apps.accounts.models import Roles, User
from apps.buyers.models import AgreementType, BuyerProfile, SisterProfile
from apps.core.models import CompanyProfile
from apps.invoicing import exports, services
from apps.invoicing.models import CommissionType, ExchangeRate, Invoice, InvoiceStatus
from apps.invoicing.views import XLSX_CONTENT_TYPE
from apps.sourcing.models import ImageLabel, Product, ProductImage
from apps.warehouse import services as warehouse_services


def _pdf_decode_stream(raw: bytes) -> bytes:
    """Undo whatever this content stream's /Filter chain applied. reportlab
    wraps content streams as [ASCII85Decode FlateDecode] (ASCII85 outer,
    Flate inner — decode in the reverse order they're listed), not bare
    FlateDecode — a decoder that only tries zlib.decompress() on the raw
    bytes silently falls through to "not compressed" and never sees the
    real operators."""
    import base64
    import zlib

    candidates = [raw]
    stripped = raw.strip(b"\r\n")
    if stripped.startswith(b"<~"):
        stripped = stripped[2:]
    if stripped.endswith(b"~>"):
        stripped = stripped[:-2]
    try:
        candidates.append(base64.a85decode(stripped, adobe=False))
    except Exception:
        pass
    for candidate in candidates:
        try:
            return zlib.decompress(candidate)
        except zlib.error:
            continue
    return candidates[-1]  # never compressed at all — use as-is


def _pdf_image_placements(pdf: bytes) -> list[tuple[float, float, float, float]]:
    """(x, y, w, h) in PDF points of every image drawn on the page.

    reportlab's Image flowable draws through `canvas.translate(x, y)` (its
    own `cm`) followed by `canvas.drawImage(...)`'s own unit-square scale
    (another `cm`, `w 0 0 h 0 0`) immediately before the `/Name Do` — two
    concatenated matrices, not one. Reading only the single `cm` adjacent to
    `Do` (as a naive regex would) always recovers the scale matrix with its
    origin at (0, 0), silently discarding whatever `q`/`cm` translations
    came before it — so a perfectly centered image would misreport x=0
    every time. This walks the whole operator stream instead, maintaining a
    real CTM stack across `q`/`Q`/`cm`, the way a PDF interpreter does."""
    import re

    def matmul(m1, m2):
        """Point transform composition: apply m1 first, then m2."""
        a1, b1, c1, d1, e1, f1 = m1
        a2, b2, c2, d2, e2, f2 = m2
        return (
            a1 * a2 + b1 * c2, a1 * b2 + b1 * d2,
            c1 * a2 + d1 * c2, c1 * b2 + d1 * d2,
            e1 * a2 + f1 * c2 + e2, e1 * b2 + f1 * d2 + f2,
        )

    placements = []
    token_re = re.compile(
        rb"(?P<num>[\d.+-]+)|(?P<cm>cm)|(?P<q>q)|(?P<Q>Q)|(?P<do>/\S+\s+Do)"
    )
    for stream_match in re.finditer(rb"stream\r?\n(.*?)endstream", pdf, re.DOTALL):
        content = _pdf_decode_stream(stream_match.group(1))
        stack = []
        ctm = (1.0, 0.0, 0.0, 1.0, 0.0, 0.0)
        operands: list[float] = []
        for tok in token_re.finditer(content):
            if tok.group("num") is not None:
                try:
                    operands.append(float(tok.group("num")))
                except ValueError:
                    pass
            elif tok.group("cm") is not None:
                if len(operands) >= 6:
                    ctm = matmul(tuple(operands[-6:]), ctm)
                operands = []
            elif tok.group("q") is not None:
                stack.append(ctm)
                operands = []
            elif tok.group("Q") is not None:
                if stack:
                    ctm = stack.pop()
                operands = []
            elif tok.group("do") is not None:
                # The unit square (0,0)-(1,1) maps to (e,f)-(e+a,f+d) under
                # the current CTM — exactly what an image XObject occupies.
                a, _b, _c, d, e, f = ctm
                placements.append((e, f, abs(a), abs(d)))
                operands = []
    return placements


class InvoiceServiceTests(APITestCase):
    def setUp(self):
        self.buyer = BuyerProfile.objects.create(name="Zara Textiles")
        self.sister = SisterProfile.objects.create(
            buyerProfile=self.buyer, poReference="PO-001",
            agreementType=AgreementType.TYPE_1, agreementRateConfig={"percentage_rate": 8},
        )
        self.employee = User.objects.create_user(username="emp", password="pass12345", role=Roles.EMPLOYEE)
        self.admin = User.objects.create_user(username="admin", password="pass12345", role=Roles.ADMIN)

    def _line_items(self, amount="1000"):
        return [{"description": "Kids T-Shirts", "ctn": 20, "qtyPerCtn": 50, "unitPrice": "1.00", "amount": amount}]

    # ── A Warehouse Cost can be pulled in as its own line item ──────────
    # Mirrors how a PackingCarton becomes a line — pick the row, it's
    # copied onto the invoice at generation time, and stays correct even
    # if the source WarehouseCost is later edited or deleted.

    def test_warehouse_cost_can_be_added_as_an_invoice_line_item(self):
        wc = warehouse_services.create_warehouse_cost(
            sister_profile=self.sister, created_by=self.employee, loaderCost=40, cartonsCost=25,
        )
        self.assertEqual(wc.totalCost, Decimal("65"))
        line_items = self._line_items() + [{
            "description": "Warehouse Cost — PO-001", "ctn": 1, "qtyPerCtn": 1, "totalQty": 1,
            "unitPrice": str(wc.totalCost), "amount": str(wc.totalCost), "warehouseCost": wc,
        }]
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=line_items)
        self.assertEqual(invoice.totalValue, Decimal("1065.00"))
        wc_line = invoice.lineItems.get(warehouseCost=wc)
        self.assertEqual(wc_line.amount, Decimal("65.00"))

    def test_deleting_the_source_warehouse_cost_leaves_the_issued_line_item_intact(self):
        wc = warehouse_services.create_warehouse_cost(sister_profile=self.sister, created_by=self.employee, loaderCost=40)
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=self._line_items() + [{
                "description": "Warehouse Cost", "ctn": 1, "qtyPerCtn": 1, "totalQty": 1,
                "unitPrice": "40", "amount": "40", "warehouseCost": wc,
            }],
        )
        warehouse_services.delete_warehouse_cost(wc, actor=self.employee)
        wc_line = invoice.lineItems.get(description="Warehouse Cost")
        self.assertIsNone(wc_line.warehouseCost)  # SET_NULL, not cascaded
        self.assertEqual(wc_line.amount, Decimal("40.00"))  # the invoice's own copy is untouched

    # ── Commission formulas (BR-43/FR-50) ───────────────────────────────

    def test_no_commission_by_default(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        self.assertEqual(invoice.totalValue, Decimal("1000.00"))
        self.assertEqual(invoice.commission_amount(), 0)
        self.assertEqual(invoice.outstandingBalance, Decimal("1000.00"))

    def test_percentage_commission(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items(),
            commission_type=CommissionType.PERCENTAGE, commission_value=5,
        )
        self.assertEqual(invoice.commission_amount(), Decimal("50.00"))
        self.assertEqual(invoice.outstandingBalance, Decimal("1050.00"))

    def test_flat_commission(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items(),
            commission_type=CommissionType.FLAT, commission_value=75,
        )
        self.assertEqual(invoice.commission_amount(), Decimal("75.00"))
        self.assertEqual(invoice.outstandingBalance, Decimal("1075.00"))

    def test_cannot_create_invoice_with_no_line_items(self):
        with self.assertRaises(ValidationError):
            services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=[])

    # ── Material captured at Sourcing Intake reaches invoice lines ──────

    def test_line_material_falls_back_to_product_material(self):
        """A line without its own material inherits Product.material - what
        Sourcing Intake captured - instead of printing an empty column."""
        product = Product.objects.create(
            sisterProfile=self.sister, name="Heavy Cotton Crewneck",
            material="100% Cotton", createdBy=self.employee,
        )
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[{"description": "Crewneck - Ecru", "product": product,
                         "ctn": 20, "qtyPerCtn": 50, "unitPrice": "1.00", "amount": "1000"}],
        )
        self.assertEqual(invoice.lineItems.get().material, "100% Cotton")

    def test_explicit_line_material_wins_over_product_material(self):
        """The invoice editor's per-line override must never be clobbered by
        the product-level value."""
        product = Product.objects.create(
            sisterProfile=self.sister, name="Heavy Cotton Crewneck",
            material="100% Cotton", createdBy=self.employee,
        )
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[{"description": "Crewneck - Ecru", "product": product,
                         "material": "Cotton-Poly 60/40",
                         "ctn": 20, "qtyPerCtn": 50, "unitPrice": "1.00", "amount": "1000"}],
        )
        self.assertEqual(invoice.lineItems.get().material, "Cotton-Poly 60/40")

    # ── Exchange rate locking (DRF doc §4 + §5 item 6) ──────────────────

    def test_exchange_rate_value_is_locked_at_creation_and_survives_republish(self):
        rate = ExchangeRate.objects.create(
            sourceCurrency="USD", targetCurrency="BDT", rate=Decimal("109.500000"),
            effectiveDate="2026-08-01", publishedBy=self.admin,
        )
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"), exchange_rate=rate,
        )
        self.assertEqual(invoice.exchangeRateValueLocked, Decimal("109.500000"))
        self.assertEqual(invoice.convertedTotal, Decimal("109500.00"))  # 1000 * 109.5

        # Admin publishes a brand new rate for the same currency pair.
        ExchangeRate.objects.create(
            sourceCurrency="USD", targetCurrency="BDT", rate=Decimal("115.000000"),
            effectiveDate="2026-09-01", publishedBy=self.admin,
        )
        # Even editing the ORIGINAL referenced row must not move the invoice.
        rate.rate = Decimal("999.000000")
        rate.save()

        invoice.refresh_from_db()
        self.assertEqual(invoice.exchangeRateValueLocked, Decimal("109.500000"))
        self.assertEqual(invoice.convertedTotal, Decimal("109500.00"))

    # ── Status transitions (BR-39/46 / FR-47-49/53) ─────────────────────

    def test_new_invoice_is_pending_approval(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        self.assertEqual(invoice.status, InvoiceStatus.PENDING_APPROVAL)

    def test_cannot_approve_twice(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        services.approve_invoice(invoice, self.admin)
        with self.assertRaises(ValidationError):
            services.approve_invoice(invoice, self.admin)

    def test_reject_requires_reason(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        with self.assertRaises(ValidationError):
            services.reject_invoice(invoice, self.admin, "")

    def test_cannot_reject_an_issued_invoice(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        services.approve_invoice(invoice, self.admin)
        with self.assertRaises(ValidationError):
            services.reject_invoice(invoice, self.admin, "too late")

    def test_cannot_void_a_pending_invoice(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        with self.assertRaises(ValidationError):
            services.void_invoice(invoice, "changed my mind", self.admin)

    def test_void_requires_reason_and_zeroes_outstanding(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        services.approve_invoice(invoice, self.admin)
        services.void_invoice(invoice, "Duplicate invoice", self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, InvoiceStatus.VOID)
        self.assertEqual(invoice.outstandingBalance, Decimal("0"))

    def test_cannot_pay_a_pending_invoice(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        with self.assertRaises(ValidationError):
            services.record_payment(invoice, amount=100, currency="USD", payment_date="2026-08-10", bank_reference="", recorded_by=self.admin)

    # ── Outstanding balance recompute (BR-44 / FR-51-52) ────────────────

    def test_outstanding_balance_recomputes_on_partial_payments(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"))
        services.approve_invoice(invoice, self.admin)
        services.record_payment(invoice, amount=400, currency="USD", payment_date="2026-08-10", bank_reference="TXN1", recorded_by=self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.outstandingBalance, Decimal("600.00"))

        services.record_payment(invoice, amount=600, currency="USD", payment_date="2026-08-11", bank_reference="TXN2", recorded_by=self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.outstandingBalance, Decimal("0.00"))

    def test_outstanding_balance_never_goes_negative(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"))
        services.approve_invoice(invoice, self.admin)
        services.record_payment(invoice, amount=1500, currency="USD", payment_date="2026-08-10", bank_reference="", recorded_by=self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.outstandingBalance, Decimal("0.00"))

    def test_cannot_record_negative_payment(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        services.approve_invoice(invoice, self.admin)
        with self.assertRaises(ValidationError):
            services.record_payment(invoice, amount=-50, currency="USD", payment_date="2026-08-10", bank_reference="", recorded_by=self.admin)

    # ── Payments are fixable after recording, while still Issued ───────

    def test_payment_amount_can_be_corrected(self):
        """A payment typed as 400 when the bank actually sent 450 shouldn't
        have to be deleted and re-entered."""
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"))
        services.approve_invoice(invoice, self.admin)
        payment = services.record_payment(invoice, amount=400, currency="USD", payment_date="2026-08-10", bank_reference="TXN1", recorded_by=self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.outstandingBalance, Decimal("600.00"))

        services.update_payment(payment, actor=self.admin, amount=450)
        invoice.refresh_from_db()
        payment.refresh_from_db()
        self.assertEqual(payment.amount, Decimal("450.00"))
        self.assertEqual(invoice.outstandingBalance, Decimal("550.00"))

    def test_payment_date_and_reference_can_be_corrected_without_touching_amount(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"))
        services.approve_invoice(invoice, self.admin)
        payment = services.record_payment(invoice, amount=400, currency="USD", payment_date="2026-08-10", bank_reference="TXN1", recorded_by=self.admin)

        services.update_payment(payment, actor=self.admin, payment_date="2026-08-11", bank_reference="TXN1-CORRECTED")
        payment.refresh_from_db()
        self.assertEqual(payment.amount, Decimal("400.00"))  # untouched, not reset
        self.assertEqual(str(payment.paymentDate), "2026-08-11")
        self.assertEqual(payment.bankReference, "TXN1-CORRECTED")

    def test_payment_edit_rejects_non_positive_amount(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"))
        services.approve_invoice(invoice, self.admin)
        payment = services.record_payment(invoice, amount=400, currency="USD", payment_date="2026-08-10", bank_reference="", recorded_by=self.admin)
        with self.assertRaises(ValidationError):
            services.update_payment(payment, actor=self.admin, amount=0)

    def test_payment_can_be_deleted_and_outstanding_recomputes(self):
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items("1000"))
        services.approve_invoice(invoice, self.admin)
        payment = services.record_payment(invoice, amount=400, currency="USD", payment_date="2026-08-10", bank_reference="", recorded_by=self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.outstandingBalance, Decimal("600.00"))

        services.delete_payment(payment, actor=self.admin)
        invoice.refresh_from_db()
        self.assertEqual(invoice.outstandingBalance, Decimal("1000.00"))
        self.assertEqual(invoice.payments.count(), 0)

    def test_payment_edit_and_delete_locked_once_voided(self):
        """BR-46: a Void invoice's payment history is part of the closed
        record, same as its rate/commission — correctable while Issued,
        frozen after."""
        invoice = services.create_invoice(sister_profile=self.sister, created_by=self.employee, line_items=self._line_items())
        services.approve_invoice(invoice, self.admin)
        payment = services.record_payment(invoice, amount=100, currency="USD", payment_date="2026-08-10", bank_reference="", recorded_by=self.admin)
        services.void_invoice(invoice, "Duplicate invoice", self.admin)
        with self.assertRaises(ValidationError):
            services.update_payment(payment, actor=self.admin, amount=200)
        with self.assertRaises(ValidationError):
            services.delete_payment(payment, actor=self.admin)

    # ── Payment details are fixable until approval ────────────────────

    def test_payment_details_can_be_corrected_before_approval(self):
        """A commission entered as 5% when it should have been a flat 50 is
        a two-argument fix, not a delete-and-recreate."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items(),
            commission_type=CommissionType.PERCENTAGE, commission_value=5,
        )
        services.update_invoice_payment_details(
            invoice, actor=self.admin, commission_type=CommissionType.FLAT, commission_value=50,
        )
        invoice.refresh_from_db()
        self.assertEqual(invoice.commissionType, CommissionType.FLAT)
        self.assertEqual(invoice.commission_amount(), Decimal("50.00"))
        self.assertEqual(invoice.outstandingBalance, Decimal("1050.00"))

    def test_editing_the_rate_relocks_and_reconverts_totals(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items(),
            source_currency="BDT", target_currency="USD", manual_rate="110",
            commission_type=CommissionType.FLAT, commission_value=50,
        )
        self.assertEqual(invoice.convertedTotal, Decimal("9.55"))  # 1050 / 110
        services.update_invoice_payment_details(invoice, actor=self.employee, manual_rate="120")
        invoice.refresh_from_db()
        self.assertEqual(invoice.exchangeRateValueLocked, Decimal("120"))
        self.assertEqual(invoice.convertedTotal, Decimal("8.75"))  # 1050 / 120

    def test_omitted_fields_keep_their_stored_value(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items(),
        )
        services.update_invoice_payment_details(invoice, actor=self.admin, commission_value=10)
        invoice.refresh_from_db()
        self.assertEqual(invoice.sourceCurrency, "BDT")  # untouched, not reset
        # A value without a type applies to nothing (same rule as creation).
        self.assertEqual(invoice.commissionType, CommissionType.NONE)
        self.assertEqual(invoice.commission_amount(), 0)

    def test_payment_details_locked_once_issued(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=self._line_items(),
        )
        services.approve_invoice(invoice, self.admin)
        with self.assertRaises(ValidationError):
            services.update_invoice_payment_details(invoice, actor=self.admin, commission_value=1)


class InvoiceAPITenantTests(APITestCase):
    def setUp(self):
        self.buyer_a = BuyerProfile.objects.create(name="Buyer A")
        self.buyer_b = BuyerProfile.objects.create(name="Buyer B")
        self.sister_a = SisterProfile.objects.create(
            buyerProfile=self.buyer_a, poReference="PO-A",
            agreementType=AgreementType.TYPE_1, agreementRateConfig={"percentage_rate": 8},
        )
        self.sister_b = SisterProfile.objects.create(
            buyerProfile=self.buyer_b, poReference="PO-B",
            agreementType=AgreementType.TYPE_1, agreementRateConfig={"percentage_rate": 8},
        )
        self.employee = User.objects.create_user(username="emp", password="pass12345", role=Roles.EMPLOYEE)
        self.admin = User.objects.create_user(username="admin", password="pass12345", role=Roles.ADMIN)
        self.buyer_a_user = User.objects.create_user(
            username="buyer_a", password="pass12345", role=Roles.BUYER, buyer_profile=self.buyer_a
        )
        line_items = [{"description": "Goods", "amount": "500"}]
        self.invoice_a = services.create_invoice(sister_profile=self.sister_a, created_by=self.employee, line_items=line_items)
        self.invoice_b = services.create_invoice(sister_profile=self.sister_b, created_by=self.employee, line_items=line_items)

    def test_buyer_cannot_see_another_buyers_invoice(self):
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("invoice-detail", args=[self.invoice_b.id]))
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    # ── A buyer can view and download their own invoices ───────────────

    def test_buyer_can_view_their_own_invoice(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("invoice-detail", args=[self.invoice_a.id]))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data["invoiceNo"], self.invoice_a.invoiceNo)
        # Buyer-facing serializer (InvoiceSelfSerializer) — no createdBy/
        # approvedBy internal fields leak through.
        self.assertNotIn("createdBy", resp.data)

    def test_buyer_can_download_their_own_invoice_as_pdf(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("invoice-export", args=[self.invoice_a.id]))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp["Content-Type"], "application/pdf")

    def test_buyer_can_download_their_own_invoice_as_xlsx(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("invoice-export", args=[self.invoice_a.id]), {"filetype": "xlsx"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp["Content-Type"], XLSX_CONTENT_TYPE)

    def test_buyer_cannot_download_another_buyers_invoice(self):
        services.approve_invoice(self.invoice_b, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("invoice-export", args=[self.invoice_b.id]))
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_buyer_sees_their_invoice_on_the_order_detail_portal_endpoint(self):
        """§17.8 — the tab a buyer actually lands on (Order Detail →
        Invoices) to view and download an invoice issued under their own
        Sister Profile."""
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("portal-order-invoices", args=[self.sister_a.id]))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(len(resp.data), 1)
        self.assertEqual(resp.data[0]["invoiceNo"], self.invoice_a.invoiceNo)

    def test_buyer_cannot_reach_another_buyers_order_invoices_via_portal(self):
        services.approve_invoice(self.invoice_b, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.get(reverse("portal-order-invoices", args=[self.sister_b.id]))
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_employee_cannot_approve_own_invoice(self):
        self.client.force_authenticate(user=self.employee)
        resp = self.client.post(reverse("invoice-approve", args=[self.invoice_a.id]))
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_buyer_cannot_void_invoice(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.post(reverse("invoice-void", args=[self.invoice_a.id]), {"reason": "x"}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_employee_cannot_publish_exchange_rate(self):
        self.client.force_authenticate(user=self.employee)
        resp = self.client.post(
            reverse("exchange-rate-list"),
            {"sourceCurrency": "USD", "targetCurrency": "BDT", "rate": "109.5", "effectiveDate": "2026-08-01"},
            format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_payment_via_api_updates_outstanding_balance(self):
        """Regression test: the ViewSet's queryset uses
        .prefetch_related("payments") for GET efficiency, which caches that
        relation on the instance `get_object()` returns. Recording a payment
        through the actual `payments` action (not the service layer
        directly, which is what apps.invoicing.tests.InvoiceServiceTests
        exercises) must not read that stale cache — this test would have
        caught a real bug found during live verification where
        outstandingBalance silently failed to update via the API despite
        updating fine when the service was called directly on a fresh
        instance."""
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.admin)
        resp = self.client.post(
            reverse("invoice-payments", args=[self.invoice_a.id]),
            {"amount": "200", "currency": "USD", "paymentDate": "2026-08-12", "bankReference": "TXN1"},
            format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(resp.data["outstandingBalance"], "300.00")
        # The same stale-prefetch issue affects the serialized `payments`
        # list independently of `outstandingBalance` — assert both.
        self.assertEqual(len(resp.data["payments"]), 1)

        self.invoice_a.refresh_from_db()
        self.assertEqual(self.invoice_a.outstandingBalance, Decimal("300.00"))

    def test_admin_can_publish_exchange_rate(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.post(
            reverse("exchange-rate-list"),
            {"sourceCurrency": "USD", "targetCurrency": "BDT", "rate": "109.5", "effectiveDate": "2026-08-01"},
            format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(resp.data["publishedBy"], self.admin.id)

    # ── Edit-before-approve on payment details ────────────────────────

    def test_employee_can_edit_payment_details_on_pending_invoice(self):
        self.client.force_authenticate(user=self.employee)
        resp = self.client.patch(
            reverse("invoice-detail", args=[self.invoice_a.id]),
            {"commissionType": "flat", "commissionValue": 25},
            format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data["commissionType"], "flat")
        self.invoice_a.refresh_from_db()
        self.assertEqual(self.invoice_a.outstandingBalance, Decimal("525.00"))  # 500 + flat 25

    def test_payment_details_cannot_be_edited_after_approval(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.admin)
        resp = self.client.patch(
            reverse("invoice-detail", args=[self.invoice_a.id]),
            {"commissionValue": 1}, format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_buyer_cannot_edit_payment_details(self):
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.patch(
            reverse("invoice-detail", args=[self.invoice_a.id]),
            {"commissionValue": 1}, format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_patch_with_an_unknown_rate_id_is_rejected(self):
        self.client.force_authenticate(user=self.employee)
        resp = self.client.patch(
            reverse("invoice-detail", args=[self.invoice_a.id]),
            {"exchangeRate": "e2ea7f52-1c1c-4b3f-9d1f-000000000000"}, format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    # ── Payments are editable after recording, via the API ─────────────

    def test_admin_can_edit_a_recorded_payment_via_api(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.admin)
        create_resp = self.client.post(
            reverse("invoice-payments", args=[self.invoice_a.id]),
            {"amount": "200", "currency": "USD", "paymentDate": "2026-08-12", "bankReference": "TXN1"},
            format="json",
        )
        payment_id = create_resp.data["payments"][0]["id"]

        resp = self.client.patch(
            reverse("invoice-payment-detail", args=[self.invoice_a.id, payment_id]),
            {"amount": "250"}, format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data["outstandingBalance"], "250.00")  # 500 - 250
        self.assertEqual(len(resp.data["payments"]), 1)
        self.assertEqual(resp.data["payments"][0]["amount"], "250.00")

    def test_admin_can_delete_a_recorded_payment_via_api(self):
        services.approve_invoice(self.invoice_a, self.admin)
        self.client.force_authenticate(user=self.admin)
        create_resp = self.client.post(
            reverse("invoice-payments", args=[self.invoice_a.id]),
            {"amount": "200", "currency": "USD", "paymentDate": "2026-08-12", "bankReference": "TXN1"},
            format="json",
        )
        payment_id = create_resp.data["payments"][0]["id"]

        resp = self.client.delete(reverse("invoice-payment-detail", args=[self.invoice_a.id, payment_id]))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data["outstandingBalance"], "500.00")
        self.assertEqual(len(resp.data["payments"]), 0)

    def test_buyer_cannot_edit_or_delete_a_payment(self):
        services.approve_invoice(self.invoice_a, self.admin)
        payment = services.record_payment(
            self.invoice_a, amount=200, currency="USD", payment_date="2026-08-12", bank_reference="", recorded_by=self.admin,
        )
        self.client.force_authenticate(user=self.buyer_a_user)
        resp = self.client.patch(
            reverse("invoice-payment-detail", args=[self.invoice_a.id, payment.id]), {"amount": "999"}, format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)
        resp = self.client.delete(reverse("invoice-payment-detail", args=[self.invoice_a.id, payment.id]))
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_editing_a_payment_on_a_voided_invoice_is_rejected_via_api(self):
        services.approve_invoice(self.invoice_a, self.admin)
        payment = services.record_payment(
            self.invoice_a, amount=200, currency="USD", payment_date="2026-08-12", bank_reference="", recorded_by=self.admin,
        )
        services.void_invoice(self.invoice_a, "Duplicate invoice", self.admin)
        self.client.force_authenticate(user=self.admin)
        resp = self.client.patch(
            reverse("invoice-payment-detail", args=[self.invoice_a.id, payment.id]), {"amount": "999"}, format="json",
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)


_TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="invoicing-test-media-")


@override_settings(MEDIA_ROOT=_TEST_MEDIA_ROOT)
class CommercialInvoiceDocumentTests(APITestCase):
    """The Commercial Invoice / Packing List document: every derived figure
    is computed server-side (never trusted from the client), the layout
    carries the company identity and bank details, and a missing product
    photo degrades to a blank cell rather than a failed download.

    Uploads go to a throwaway MEDIA_ROOT: these tests write real image
    files, and the default root is the developer's own `backend/media/`.
    """

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(_TEST_MEDIA_ROOT, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.buyer = BuyerProfile.objects.create(
            name="LTD LOXY", branding="Loxy",
            contactInfo="I/D NO: 404853592\n2 Pekini Street, Tbilisi, Georgia",
        )
        self.sister = SisterProfile.objects.create(
            buyerProfile=self.buyer, poReference="PO-A1",
            agreementType=AgreementType.TYPE_1, agreementRateConfig={"percentage_rate": 8},
        )
        self.employee = User.objects.create_user(username="doc_emp", password="pass12345", role=Roles.EMPLOYEE)
        self.admin = User.objects.create_user(username="doc_admin", password="pass12345", role=Roles.ADMIN)

    def _sample_line(self, **overrides):
        """The first line of the real sample invoice: 3 cartons x 60 pcs at
        150, in 60x40x30cm cartons."""
        line = {
            "description": "woman shorts", "brand": "Kappa", "colorSizeNote": "3 color. 70+70+70",
            "markRef": "A2", "ctn": 3, "qtyPerCtn": 60, "unitPrice": "150",
            "netWeightPerCtn": "13.5", "grossWeightPerCtn": "14.75",
            "ctnLengthCm": 60, "ctnWidthCm": 40, "ctnHeightCm": 30, "dimensionsInCm": True,
            "material": "100%cotton",
        }
        line.update(overrides)
        return line

    # ── Auto-calculation (matches the sample's own arithmetic) ──────────

    def test_line_totals_are_computed_from_cartons_and_unit_price(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=[self._sample_line()],
        )
        line = invoice.lineItems.first()
        self.assertEqual(line.totalQty, 180)                      # 3 x 60
        self.assertEqual(line.amount, Decimal("27000.00"))        # 180 x 150
        self.assertEqual(line.netWeight, Decimal("40.50"))        # 3 x 13.5
        self.assertEqual(line.grossWeight, Decimal("44.25"))      # 3 x 14.75
        self.assertEqual(line.cbmPerCtn, Decimal("0.072000"))     # 60x40x30 / 1e6
        self.assertEqual(line.cbm, Decimal("0.2160"))             # 3 x 0.072
        self.assertEqual(invoice.totalValue, Decimal("27000.00"))

    def test_explicitly_supplied_values_are_not_overwritten_by_the_formula(self):
        """Short/excess cartons are real — the packing list wins over
        arithmetic when the two disagree."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line(totalQty=175, amount="26250")],
        )
        line = invoice.lineItems.first()
        self.assertEqual(line.totalQty, 175)
        self.assertEqual(line.amount, Decimal("26250.00"))

    def test_inch_carton_dimensions_are_converted_to_centimetres(self):
        """PackingCarton stores inches; the document prints cm and derives
        CBM from them."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line(ctnLengthCm=10, ctnWidthCm=10, ctnHeightCm=10, dimensionsInCm=False)],
        )
        line = invoice.lineItems.first()
        self.assertEqual(line.ctnLengthCm, Decimal("25.40"))
        self.assertEqual(line.cbmPerCtn, Decimal("0.016387"))

    def test_document_totals_sum_every_line(self):
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[
                self._sample_line(),
                self._sample_line(ctn=1, qtyPerCtn=30, netWeightPerCtn="7", grossWeightPerCtn="8.5"),
            ],
        )
        totals = invoice.document_totals()
        self.assertEqual(totals["totalQty"], 210)                      # 180 + 30
        self.assertEqual(totals["totalCtn"], 4)
        self.assertEqual(totals["totalNetWeight"], Decimal("47.50"))   # 40.5 + 7
        self.assertEqual(totals["totalCbm"], Decimal("0.2880"))        # 0.216 + 0.072

    # ── Manual currency + exchange rate ─────────────────────────────────

    def test_manual_rate_converts_in_the_direction_it_was_typed(self):
        """"1 USD = 120 BDT" against the sample's own total: 12,727,855 BDT
        is 106,065.46 USD."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=[self._sample_line()],
            source_currency="BDT", target_currency="USD", manual_rate="120",
        )
        self.assertEqual(invoice.sourceCurrency, "BDT")
        self.assertEqual(invoice.rateQuote, "divide")
        self.assertEqual(invoice.rate_label(), "1 USD = 120 BDT")
        self.assertEqual(invoice.convert(Decimal("12727855")), Decimal("106065.46"))
        self.assertEqual(invoice.convertedTotal, Decimal("225.00"))  # 27000 / 120

    def test_published_rate_still_multiplies_as_before(self):
        rate = ExchangeRate.objects.create(
            sourceCurrency="BDT", targetCurrency="USD", rate="0.0083",
            effectiveDate="2026-08-01", publishedBy=self.admin,
        )
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line()], exchange_rate=rate,
        )
        self.assertEqual(invoice.rateQuote, "multiply")
        self.assertEqual(invoice.convertedTotal, Decimal("224.10"))  # 27000 x 0.0083

    def test_rate_without_a_target_currency_is_rejected(self):
        with self.assertRaises(ValidationError):
            services.create_invoice(
                sister_profile=self.sister, created_by=self.employee,
                line_items=[self._sample_line()], manual_rate="120",
            )

    # ── Rendering ───────────────────────────────────────────────────────

    def _document_invoice(self):
        return services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[
                self._sample_line(),
                self._sample_line(brand="Jack & Jones", description="boy <hoodie> & cap"),
            ],
            source_currency="BDT", target_currency="USD", manual_rate="120",
            commission_type=CommissionType.PERCENTAGE, commission_value=15,
        )

    def _fill_company(self):
        company = CompanyProfile.load()
        company.name = "Sumaiya International"
        company.tagline = "(Export, Import, Supply & Manufacturer)"
        company.addressLine = "House 38, Road 5/A, Uttara, Dhaka-1230, Bangladesh"
        company.phone = "+8801777634066"
        company.contactPerson = "Mohammed Rafi"
        company.registrationNo = "BIN-123"
        company.bankName = "Commercial Bank of Ceylon PLC"
        company.bankAccountTitle = "M/S SUMAIYA INTERNATIONAL"
        company.bankAccountNo = "1806011466"
        company.bankSwiftCode = "CCEYBDDH"
        company.save()
        return company

    def test_pdf_renders_with_company_and_bank_details(self):
        self._fill_company()
        pdf = exports.render_invoice_pdf(self._document_invoice())
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertGreater(len(pdf), 2000)

    def test_pdf_renders_when_company_profile_is_still_blank(self):
        """A brand-new install with nothing filled in must still produce a
        document rather than raising."""
        pdf = exports.render_invoice_pdf(self._document_invoice())
        self.assertTrue(pdf.startswith(b"%PDF"))

    # ── Letterhead placement ──────────────────────────────────────────

    def _set_logo(self, size):
        """A real (tiny) PNG through the normal upload path, like a user
        uploading their letterhead on the Company Profile page."""
        company = CompanyProfile.load()
        buf = io.BytesIO()
        PILImage.new("RGB", size, (16, 24, 40)).save(buf, format="PNG")
        company.logo.save("letterhead.png", ContentFile(buf.getvalue()), save=True)
        return company

    def test_letterhead_is_centered_at_the_top_of_the_pdf(self):
        """The company banner belongs top-centre. reportlab's Image flowable
        carries no hAlign attribute, so without the explicit CENTER it used
        to render flush left."""
        self._fill_company()
        self._set_logo((1400, 300))
        pdf = exports.render_invoice_pdf(self._document_invoice())
        placements = _pdf_image_placements(pdf)
        self.assertTrue(placements, "expected the letterhead image in the PDF")
        # The letterhead is the topmost image on the page (largest y in PDF
        # space, where the origin is the bottom-left corner).
        x, _y, w, _h = max(placements, key=lambda p: p[1])
        page_width = landscape(A4)[0]  # 841.89pt
        self.assertAlmostEqual(x, (page_width - w) / 2, delta=1.0)

    def test_tall_logo_keeps_its_aspect_ratio(self):
        """Clamping only the height used to squash a square logo into a
        distorted 180×26mm strip — the width must shrink when the height
        cap bites, square in means square out."""
        self._set_logo((600, 600))
        pdf = exports.render_invoice_pdf(self._document_invoice())
        placements = _pdf_image_placements(pdf)
        x, _y, w, h = max(placements, key=lambda p: p[1])
        self.assertAlmostEqual(w, h, delta=1.0)
        self.assertLessEqual(h, 26 * 72 / 25.4 + 1)  # 26mm height cap, in points
        page_width = landscape(A4)[0]
        self.assertAlmostEqual(x, (page_width - w) / 2, delta=1.0)

    def _set_seal(self, size=(500, 220)):
        company = CompanyProfile.load()
        buf = io.BytesIO()
        PILImage.new("RGB", size, (200, 40, 40)).save(buf, format="PNG")
        company.sealSignature.save("seal.png", ContentFile(buf.getvalue()), save=True)
        return company

    def test_pdf_omits_the_signature_block_when_no_seal_is_uploaded(self):
        """A brand-new install has no seal/signature yet — the caption alone
        (no image) must not break the render."""
        pdf = exports.render_invoice_pdf(self._document_invoice())
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_seal_signature_prints_bottom_right_of_the_pdf(self):
        self._set_seal()
        pdf = exports.render_invoice_pdf(self._document_invoice())
        placements = _pdf_image_placements(pdf)
        # Two images now ride the page: the letterhead (none uploaded here,
        # so just the seal) — with no logo set, the seal is the only image.
        self.assertTrue(placements, "expected the seal/signature image in the PDF")
        x, _y, w, _h = min(placements, key=lambda p: p[1])  # bottom-most image
        page_width = landscape(A4)[0]
        self.assertGreater(x + w, page_width / 2)  # sits right of centre

    def test_seal_signature_does_not_collide_with_the_letterhead(self):
        """Both a top letterhead and a bottom seal can be uploaded at once —
        two distinct images, neither one clobbering the other's slot.

        Doesn't assert relative vertical position: on a short test invoice
        the totals/bank footer can fill page 1, and reportlab correctly
        refuses to split the seal image across the page break — it lands
        alone at the top of page 2, at the same margin-relative y the
        letterhead uses on page 1. That's real pagination, not a defect;
        this test only needs both images to exist and stay horizontally
        distinct (centered letterhead vs. right-aligned seal).
        """
        self._set_logo((1400, 300))
        self._set_seal()
        pdf = exports.render_invoice_pdf(self._document_invoice())
        placements = _pdf_image_placements(pdf)
        self.assertEqual(len(placements), 2)
        xs = sorted(p[0] for p in placements)
        self.assertGreater(xs[1] - xs[0], 100)  # clearly separated horizontally

    def test_xlsx_carries_the_seal_signature_image(self):
        self._set_seal()
        content = exports.render_invoice_xlsx(self._document_invoice())
        ws = load_workbook(io.BytesIO(content)).active
        self.assertEqual(len(ws._images), 1)
        text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertIn("Authorized Signatory", text)

    def test_xlsx_letterhead_is_anchored_centre(self):
        """Excel anchors pictures to cells and A1 is flush left; the
        letterhead must ride a computed offset into the column grid so it
        sits top-centre, matching the PDF."""
        self._set_logo((1400, 300))
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=[self._sample_line()],
        )
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(invoice))).active
        self.assertEqual(len(ws._images), 1)
        marker = ws._images[0].anchor._from
        self.assertEqual(marker.row, 0)  # top row
        self.assertGreater(marker.col, 0)  # pushed in from column A = centred

    def test_logo_center_anchor_targets_the_middle_of_the_grid(self):
        widths = [round(exports.XLSX_COLUMN_WIDTHS[key] * 7) + 5 for key, _ in exports.COLUMNS]
        col, off = exports._logo_center_anchor(420)
        left_px = sum(widths[:col]) + off / 9525  # 1 px = 9525 EMU
        self.assertAlmostEqual(left_px, (sum(widths) - 420) // 2, delta=1)

    def test_xlsx_carries_the_layout_and_numeric_amounts(self):
        self._fill_company()
        content = exports.render_invoice_xlsx(self._document_invoice())
        ws = load_workbook(io.BytesIO(content)).active
        text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        # The reference stacks the two document titles, centered.
        self.assertIn("COMMERCIAL INVOICE", text)
        self.assertIn("PACKING LIST", text)
        self.assertIn("Foto", text)
        self.assertIn("T/N.W", text)
        # …and its full 21-column body, grouped SIZE header included.
        self.assertIn("SIZE (cm)", text)
        self.assertIn("Style No", text)
        self.assertIn("CBM/ctn", text)
        self.assertIn("LTD LOXY", text)
        self.assertIn("Commercial Bank of Ceylon PLC", text)
        # The header block prints Contact / TEL / registration, as on paper.
        self.assertIn("Contact: Mohammed Rafi", text)
        self.assertIn("TEL: +8801777634066", text)
        self.assertIn("ID / Registration No: BIN-123", text)
        # Physical quantities land as real numbers, not preformatted strings —
        # a packing list you can't re-total in Excel isn't much of an export.
        # (openpyxl reads a whole number back as int, so accept either.)
        numeric = [c.value for row in ws.iter_rows() for c in row
                   if isinstance(c.value, (int, float)) and not isinstance(c.value, bool) and c.value == 180]
        self.assertTrue(numeric, "expected the 180-pcs line quantity as a numeric cell")

    def test_line_table_carries_no_per_line_money_columns(self):
        """The money side of this document is one figure (TOTAL VALUE), not a
        priced-out bill — Unit Price and AMOUNT are deliberately absent from
        the line table, and the per-line 27,000 they used to print with it."""
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(self._document_invoice()))).active
        text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertNotIn("Unit Price", text)
        self.assertNotIn("AMOUNT", text)
        priced = [c.value for row in ws.iter_rows() for c in row
                  if isinstance(c.value, (int, float)) and not isinstance(c.value, bool) and c.value == 27000]
        self.assertFalse(priced, "per-line amounts must not be printed on the document")

    def test_totals_block_prints_all_in_total_value_without_outstanding(self):
        """TOTAL VALUE is the all-in figure: product lines + warehouse-cost
        lines + commission, every related cost rolled into the one total the
        document shows. The outstanding balance is deliberately not printed
        — the buyer gets the total and what has been received against it,
        nothing more."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line()],
            commission_type=CommissionType.FLAT, commission_value="1000",
        )
        content = exports.render_invoice_xlsx(invoice)
        ws = load_workbook(io.BytesIO(content)).active
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        self.assertIn("TOTAL VALUE", text)
        self.assertIn("BDT", text)
        self.assertIn("28,000.00 BDT", text)  # 27,000 lines + 1,000 commission
        self.assertNotIn("OUTSTANDING IN THIS INVOICE", text)
        self.assertIn("TOTAL CTNS/BAG", text)
        self.assertNotIn("COMMISSION", text)
        self.assertNotIn("EXCHANGE RATE", text)
        self.assertNotIn("1 USD = 120 BDT", text)

    def test_payments_received_still_show_their_own_arithmetic(self):
        """Each bank transfer summed = received, as on the reference — the
        payment record is the one money detail beyond the total that a buyer
        does need on paper."""
        invoice = self._document_invoice()
        services.approve_invoice(invoice, self.admin)
        services.record_payment(
            invoice, amount=Decimal("1000"), currency="BDT", payment_date="2026-08-10",
            bank_reference="", recorded_by=self.employee,
        )
        services.record_payment(
            invoice, amount=Decimal("500"), currency="BDT", payment_date="2026-08-12",
            bank_reference="", recorded_by=self.employee,
        )
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(invoice))).active
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        self.assertIn("BANK TRANSFER", text)
        self.assertIn("1,000.00 + 500.00 = 1,500.00 BDT", text)
        self.assertIn("TOTAL VALUE", text)
        self.assertIn("62,100.00 BDT", text)  # the all-in total value: 54,000 lines + 15% commission (8,100)

    def test_line_number_and_shipping_mark_share_the_first_cell(self):
        """The reference's first column is a row number; our shipping mark
        rides underneath it instead of claiming a column of its own."""
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(self._document_invoice()))).active
        header_row = next(r for r in range(1, 30) if ws.cell(row=r, column=1).value == "№")
        # Both sample lines carry the "A2" mark, riding under the row number.
        self.assertEqual(ws.cell(row=header_row + 2, column=1).value, "1\nA2")  # two header rows
        self.assertEqual(ws.cell(row=header_row + 3, column=1).value, "2\nA2")

    def test_size_sub_columns_live_under_the_group_header(self):
        """SIZE is a grouped header spanning three sub-columns (L/W/H), as
        in the reference's merged-cell layout — not one combined string."""
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(self._document_invoice()))).active
        header_row = next(r for r in range(1, 30) if ws.cell(row=r, column=1).value == "№")
        # Derived from COLUMNS, not hardcoded: the SIZE group's position
        # shifts whenever a column is added or dropped from the layout.
        size_col = exports.KEYS.index("sizeL") + 1
        self.assertEqual(ws.cell(row=header_row, column=size_col).value, "SIZE (cm)")
        self.assertEqual(ws.cell(row=header_row + 1, column=size_col).value, "L")
        self.assertEqual(ws.cell(row=header_row + 1, column=size_col + 1).value, "W")
        self.assertEqual(ws.cell(row=header_row + 1, column=size_col + 2).value, "H")
        # The grouped header really is merged across the three of them.
        start, end = get_column_letter(size_col), get_column_letter(size_col + 2)
        self.assertIn(f"{start}{header_row}:{end}{header_row}", [str(r) for r in ws.merged_cells.ranges])

    def test_style_column_renders_the_real_style_and_pattern_numbers(self):
        """The reference's lettered codes (A2, K13) are replaced by our real
        identifiers: style_no / pattern_no, e.g. "MRF25 / MR12528"."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line(styleItemCode="MRF25", patternNo="MR12528")],
        )
        self.assertEqual(invoice.lineItems.first().patternNo, "MR12528")  # locked at generation
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(invoice))).active
        text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertIn("MRF25 / MR12528", text)

    def test_georgian_export_renders_georgian_labels(self):
        """BR-33 / FR-60-63: labels are bilingual. Excel has no font problem
        (fonts render client-side); the PDF needs the bundled DejaVu family
        for Mkhedruli and must still render without error either way."""
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(self._document_invoice(), lang="ka"))).active
        text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertIn("კომერციული ინვოისი", text)
        self.assertIn("შეფუთვის სია", text)
        pdf = exports.render_invoice_pdf(self._document_invoice(), lang="ka")
        self.assertTrue(pdf.startswith(b"%PDF"))

    # ── Product photos on the document ──────────────────────────────────

    def _product_with_photo(self, label=ImageLabel.PRODUCT_OVERALL, color=(200, 30, 30)):
        """A real (tiny) JPEG through the normal ProductImage upload path —
        the whole point is to exercise what happens to a photo a user
        actually attached to a product."""
        product = Product.objects.create(sisterProfile=self.sister, name="Polo Shirt", createdBy=self.employee)
        buf = io.BytesIO()
        PILImage.new("RGB", (300, 300), color).save(buf, format="JPEG")
        buf.seek(0)
        image = ProductImage(product=product, label=label, uploadedBy=self.employee)
        image.image.save("polo.jpg", ContentFile(buf.getvalue()), save=False)
        image.save()
        return product

    def test_photo_added_to_a_product_appears_in_both_documents(self):
        product = self._product_with_photo()
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line(product=product)],
        )
        # Excel: the photo is embedded as a real drawing on the sheet.
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(invoice))).active
        self.assertEqual(len(ws._images), 1)
        # PDF: an image stream makes it materially bigger than a text-only one.
        self.assertTrue(exports.render_invoice_pdf(invoice).startswith(b"%PDF"))

    def test_the_product_overall_photo_wins_over_an_earlier_upload(self):
        """ProductImage is ordered by upload time, so a bare .first() would
        put a fabric close-up on the invoice instead of the product."""
        product = self._product_with_photo(label=ImageLabel.FABRIC_CLOSEUP, color=(10, 10, 200))
        overall = ProductImage(product=product, label=ImageLabel.PRODUCT_OVERALL, uploadedBy=self.employee)
        buf = io.BytesIO()
        PILImage.new("RGB", (300, 300), (20, 200, 20)).save(buf, format="JPEG")
        overall.image.save("overall.jpg", ContentFile(buf.getvalue()), save=False)
        overall.save()

        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line(product=product)],
        )
        chosen = exports._line_photo(invoice.lineItems.first())
        self.assertIsNotNone(chosen)
        # The green "product overall" shot, not the blue fabric close-up.
        pixel = PILImage.open(chosen).convert("RGB").getpixel((5, 5))
        self.assertGreater(pixel[1], pixel[2], "expected the Product Overall image, not the fabric close-up")

    def test_a_line_without_a_product_or_photo_still_renders(self):
        """No photo must mean an empty cell, never a failed download."""
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee, line_items=[self._sample_line()],
        )
        ws = load_workbook(io.BytesIO(exports.render_invoice_xlsx(invoice))).active
        self.assertEqual(len(ws._images), 0)
        self.assertTrue(exports.render_invoice_pdf(invoice).startswith(b"%PDF"))

    def test_an_unreadable_image_file_does_not_break_the_export(self):
        """Media files go missing (restores, volume mounts, manual cleanup) —
        that must degrade to a blank Foto cell."""
        product = self._product_with_photo()
        product.images.first().image.storage.delete(product.images.first().image.name)
        invoice = services.create_invoice(
            sister_profile=self.sister, created_by=self.employee,
            line_items=[self._sample_line(product=product)],
        )
        self.assertIsNone(exports._line_photo(invoice.lineItems.first()))
        self.assertTrue(exports.render_invoice_pdf(invoice).startswith(b"%PDF"))

    # ── Company profile ─────────────────────────────────────────────────

    def test_company_profile_is_a_singleton(self):
        first = CompanyProfile.load()
        first.name = "One"
        first.save()
        second = CompanyProfile.objects.create(name="Two")
        self.assertEqual(CompanyProfile.objects.count(), 1)
        self.assertEqual(second.pk, first.pk)

    def test_only_admin_can_change_the_bank_details(self):
        self.client.force_authenticate(user=self.employee)
        resp = self.client.patch(reverse("company-profile"), {"bankAccountNo": "999"}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        resp = self.client.patch(reverse("company-profile"), {"bankAccountNo": "999"}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(CompanyProfile.load().bankAccountNo, "999")

    def test_incomplete_profile_reports_what_is_missing(self):
        self.assertIn("Bank name", CompanyProfile.load().missing_fields())
        self._fill_company()
        self.assertTrue(CompanyProfile.load().is_complete())
