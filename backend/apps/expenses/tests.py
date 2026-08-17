import csv
import io
from datetime import datetime, timezone as dt_timezone
from decimal import Decimal
from unittest.mock import patch

from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from apps.accounts.models import Roles, User
from apps.audit.models import AuditLogEntry
from apps.buyers.models import AgreementType, BuyerProfile, SisterProfile
from apps.expenses import exports
from apps.expenses.models import Expense, SourceType
from apps.expenses.services import record_expense
from apps.sourcing.models import Product

EXPORT_URL = reverse("expense-export")


class ExpenseExportTests(APITestCase):
    """FR-72-73 download side: the file must contain exactly the filtered
    slice the caller asked for, never blend currencies, and never be
    reachable by a Buyer-role user."""

    def setUp(self):
        self.admin = User.objects.create_user(username="admin", password="pass12345", role=Roles.ADMIN)
        self.rep = User.objects.create_user(username="rep", password="pass12345", role=Roles.COMPANY_REP)

        self.buyer_a = BuyerProfile.objects.create(name="Alpha Textiles")
        self.buyer_b = BuyerProfile.objects.create(name="Beta Garments")
        self.sister_a1 = self._sister(self.buyer_a, "PO-A1")
        self.sister_a2 = self._sister(self.buyer_a, "PO-A2")
        self.sister_b1 = self._sister(self.buyer_b, "PO-B1")

        self.product = Product.objects.create(sisterProfile=self.sister_a1, name="Denim Shirt", createdBy=self.rep)

        # Buyer A / PO-A1 — two currencies on purpose.
        self.e1 = record_expense(
            sister_profile=self.sister_a1, source_type=SourceType.SOURCING_ADVANCE, amount=1000,
            created_by=self.admin, product=self.product, remarks="Advance for sourcing trip",
        )
        self.e2 = record_expense(
            sister_profile=self.sister_a1, source_type=SourceType.QC_LUNCH, amount=250,
            created_by=self.rep, currency="USD",
        )
        # Buyer A / PO-A2
        self.e3 = record_expense(
            sister_profile=self.sister_a2, source_type=SourceType.WAREHOUSE_LOADER, amount=400,
            created_by=self.admin,
        )
        # Buyer B
        self.e4 = record_expense(
            sister_profile=self.sister_b1, source_type=SourceType.QC_CARRYING, amount=700,
            created_by=self.admin,
        )

        # createdAt is auto_now_add, so backdate through the queryset to get
        # a testable date range.
        self._set_created(self.e1, datetime(2026, 1, 10, 9, 0, tzinfo=dt_timezone.utc))
        self._set_created(self.e2, datetime(2026, 2, 15, 9, 0, tzinfo=dt_timezone.utc))
        self._set_created(self.e3, datetime(2026, 3, 20, 9, 0, tzinfo=dt_timezone.utc))
        self._set_created(self.e4, datetime(2026, 3, 25, 9, 0, tzinfo=dt_timezone.utc))

    def _sister(self, buyer, po_reference):
        return SisterProfile.objects.create(
            buyerProfile=buyer, poReference=po_reference,
            agreementType=AgreementType.TYPE_1,
        )

    def _set_created(self, expense, when):
        Expense.objects.filter(pk=expense.pk).update(createdAt=when)

    def _csv_rows(self, response):
        text = response.content.decode("utf-8-sig")
        return list(csv.reader(io.StringIO(text)))

    # ── Access ──────────────────────────────────────────────────────────

    def test_buyer_role_cannot_export(self):
        """The buyer-facing serializer withholds the internal columns every
        export carries, so the download stays supplier-side only."""
        buyer_user = User.objects.create_user(
            username="buyer_a", password="pass12345", role=Roles.BUYER, buyer_profile=self.buyer_a,
        )
        self.client.force_authenticate(user=buyer_user)
        resp = self.client.get(EXPORT_URL)
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_unauthenticated_cannot_export(self):
        resp = self.client.get(EXPORT_URL)
        self.assertIn(resp.status_code, (status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN))

    # ── Filters ─────────────────────────────────────────────────────────

    def test_buyer_profile_filter_limits_rows_to_that_buyer(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"filetype": "csv", "buyerProfile": str(self.buyer_a.id)})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        rows = self._csv_rows(resp)
        self.assertEqual(len(rows) - 1, 3)  # e1, e2, e3 — not buyer B's e4
        self.assertTrue(all(r[1] == "Alpha Textiles" for r in rows[1:]))

    def test_sister_profile_filter_limits_rows_to_that_po(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"filetype": "csv", "sisterProfile": str(self.sister_a1.id)})
        rows = self._csv_rows(resp)
        self.assertEqual(len(rows) - 1, 2)
        self.assertTrue(all(r[2] == "PO-A1" for r in rows[1:]))

    def test_date_range_filter_is_inclusive_on_both_bounds(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {
            "filetype": "csv", "dateFrom": "2026-02-15", "dateTo": "2026-03-20",
        })
        rows = self._csv_rows(resp)
        dates = sorted(r[0][:10] for r in rows[1:])
        self.assertEqual(dates, ["2026-02-15", "2026-03-20"])

    def test_source_type_accepts_a_comma_separated_list(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {
            "filetype": "csv", "sourceType": f"{SourceType.QC_LUNCH},{SourceType.QC_CARRYING}",
        })
        rows = self._csv_rows(resp)
        self.assertEqual(sorted(r[5] for r in rows[1:]), ["QC Carrying", "QC Lunch"])

    def test_search_matches_remarks(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"filetype": "csv", "search": "sourcing trip"})
        rows = self._csv_rows(resp)
        self.assertEqual(len(rows) - 1, 1)

    def test_combined_filters_narrow_to_a_single_row(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {
            "filetype": "csv", "buyerProfile": str(self.buyer_a.id),
            "sisterProfile": str(self.sister_a1.id), "currency": "USD",
            "dateFrom": "2026-01-01", "dateTo": "2026-12-31",
        })
        rows = self._csv_rows(resp)
        self.assertEqual(len(rows) - 1, 1)
        self.assertEqual(rows[1][8], "USD")

    def test_list_endpoint_honours_the_same_filters_as_the_export(self):
        """The download promises "exactly what the table shows" — which only
        holds while both read the same query params."""
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(reverse("expense-list"), {"buyerProfile": str(self.buyer_a.id)})
        self.assertEqual(resp.data["count"], 3)

    # ── Formats ─────────────────────────────────────────────────────────

    def test_xlsx_export_has_summary_and_detail_sheets_and_filename(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {
            "buyerProfile": str(self.buyer_a.id), "dateFrom": "2026-01-01", "dateTo": "2026-03-31",
        })
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn("expenses-alpha-textiles-2026-01-01_to_2026-03-31.xlsx", resp["Content-Disposition"])

        wb = load_workbook(io.BytesIO(resp.content))
        self.assertEqual(wb.sheetnames, ["Summary", "Expenses"])
        summary_text = "\n".join(
            str(c.value) for row in wb["Summary"].iter_rows() for c in row if c.value is not None
        )
        self.assertIn("Alpha Textiles", summary_text)
        self.assertIn("2026-01-01 to 2026-03-31", summary_text)

    def test_pdf_export_returns_a_pdf(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"filetype": "pdf", "groupBy": "buyer"})
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_pdf_survives_ampersands_and_angle_brackets_in_user_text(self):
        """reportlab parses Paragraph text as mini-HTML, so an unescaped
        "R&D" in a remark would abort the whole render."""
        record_expense(
            sister_profile=self.sister_a1, source_type=SourceType.EXTRA_COST, amount=90,
            created_by=self.admin, remarks="R&D sample <urgent>", field_name="Smith & Co",
        )
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {
            "filetype": "pdf", "groupBy": "sourceType", "search": "R&D <sample>",
        })
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_csv_carries_a_bom_so_excel_reads_utf8(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"filetype": "csv"})
        self.assertTrue(resp.content.startswith(b"\xef\xbb\xbf"))

    def test_unknown_filetype_and_group_by_are_rejected(self):
        self.client.force_authenticate(user=self.admin)
        self.assertEqual(
            self.client.get(EXPORT_URL, {"filetype": "docx"}).status_code, status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(
            self.client.get(EXPORT_URL, {"groupBy": "phase-of-the-moon"}).status_code,
            status.HTTP_400_BAD_REQUEST,
        )

    # ── Grouping and totals ─────────────────────────────────────────────

    def test_grouped_csv_prepends_the_group_column(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"filetype": "csv", "groupBy": "buyer"})
        rows = self._csv_rows(resp)
        self.assertEqual(rows[0][0], "Group (Buyer)")
        self.assertEqual(sorted({r[0] for r in rows[1:]}), ["Alpha Textiles", "Beta Garments"])

    def test_totals_are_per_currency_never_blended(self):
        """PO-A1 holds 1000 BDT and 250 USD — two summary rows, not one
        1250 of nothing."""
        expenses = list(Expense.objects.filter(sisterProfile=self.sister_a1))
        rows = exports.summary_rows(expenses, "sisterProfile")
        self.assertEqual(
            sorted((currency, total) for _, currency, _, total in rows),
            [("BDT", Decimal("1000.00")), ("USD", Decimal("250.00"))],
        )
        self.assertEqual(exports.totals_line(expenses), "1,000.00 BDT  +  250.00 USD")

    def test_grouped_xlsx_writes_a_subtotal_row_per_group_and_currency(self):
        self.client.force_authenticate(user=self.admin)
        resp = self.client.get(EXPORT_URL, {"groupBy": "sisterProfile"})
        wb = load_workbook(io.BytesIO(resp.content))
        subtotals = [
            str(c.value) for row in wb["Expenses"].iter_rows(min_col=1, max_col=1)
            for c in row if c.value and "subtotal" in str(c.value)
        ]
        self.assertIn("PO-A1 — subtotal (BDT)", subtotals)
        self.assertIn("PO-A1 — subtotal (USD)", subtotals)
        self.assertIn("PO-A2 — subtotal (BDT)", subtotals)

    # ── Guardrails ──────────────────────────────────────────────────────

    def test_oversized_export_is_refused_with_a_narrow_your_filters_message(self):
        self.client.force_authenticate(user=self.admin)
        with patch.object(exports, "MAX_EXPORT_ROWS", 2):
            resp = self.client.get(EXPORT_URL, {"filetype": "csv"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("export limit", str(resp.data))

    def test_export_is_written_to_the_audit_log(self):
        self.client.force_authenticate(user=self.admin)
        self.client.get(EXPORT_URL, {"filetype": "csv", "buyerProfile": str(self.buyer_a.id)})
        entry = AuditLogEntry.objects.filter(action="EXPORT_EXPENSES").first()
        self.assertIsNotNone(entry)
        self.assertEqual(entry.actor, self.admin)
        self.assertEqual(entry.afterSnapshot["rows"], 3)
        self.assertEqual(entry.afterSnapshot["filters"]["buyerProfile"], str(self.buyer_a.id))
