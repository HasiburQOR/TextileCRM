"""
Central Expense Table exports (Excel + CSV + PDF) — the download side of
BR-48 / FR-72-73. Nothing here computes or re-derives a cost: it only
formats Expense rows the services already wrote, in whatever slice the
caller filtered down to (buyer, sister profile / PO, date range, source
type, …) and optionally grouped.

Two rules the layout follows everywhere:

  * Amounts are NEVER summed across currencies. Every total — group
    subtotal and grand total alike — is per-currency, because an Expense
    carries its own `currency` and BDT + USD is not a number.
  * The filter block that produced the file is written into the file
    itself (xlsx/pdf) and into its filename, so a downloaded copy is
    still self-describing a month later when nobody remembers which
    filters were on screen.
"""

import csv
import io
from collections import OrderedDict
from decimal import Decimal
from xml.sax.saxutils import escape

from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

HEADER_FILL = "1E2937"
LINE_HEADER_FILL = "E5E7EB"
GROUP_FILL = "F1F5F9"
LINE_HEADER_FILL_HEX = "#" + LINE_HEADER_FILL

# xlsx/pdf builders hold every row in memory (openpyxl and reportlab both
# do), so an unbounded "download everything" on a multi-year table is a
# memory incident waiting to happen. The view turns this into a "narrow
# your filters" 400 rather than trying and dying.
MAX_EXPORT_ROWS = 20000

DETAIL_HEADERS = [
    "Date", "Buyer", "Sister Profile / PO", "Product", "Style No",
    "Source Type", "Field", "Amount", "Currency", "Remarks", "Recorded By",
]

# groupBy value -> (column header used for the group, label extractor).
# "none" is handled separately: no grouping column, no subtotals.
GROUP_BY_OPTIONS = OrderedDict([
    ("none", ("", None)),
    ("buyer", ("Buyer", lambda e: e.sisterProfile.buyerProfile.name)),
    ("sisterProfile", ("Sister Profile / PO", lambda e: _sister_label(e.sisterProfile))),
    ("product", ("Product", lambda e: e.product.name if e.product else "— No product —")),
    ("sourceType", ("Source Type", lambda e: e.get_sourceType_display())),
    ("recordedBy", ("Recorded By", lambda e: e.createdBy.display_name() if e.createdBy else "—")),
    ("month", ("Month", lambda e: e.createdAt.strftime("%Y-%m"))),
    ("currency", ("Currency", lambda e: e.currency)),
])


def _money(value) -> str:
    return f"{float(value or 0):,.2f}"


def _esc(value) -> str:
    """reportlab's Paragraph parses its text as mini-HTML, so any `&` or
    `<` coming out of the database — a remark reading "R&D sample", a
    buyer named "Smith & Co" — raises a parse error mid-render and takes
    the whole download with it. Everything user-entered goes through here
    before it reaches a Paragraph."""
    return escape(str(value if value is not None else ""))


def _sister_label(sister_profile) -> str:
    return sister_profile.poReference or sister_profile.referenceCode or str(sister_profile.id)


def group_label(expense, group_by: str) -> str:
    extractor = GROUP_BY_OPTIONS.get(group_by, ("", None))[1]
    return extractor(expense) if extractor else ""


def _detail_row(expense) -> list:
    """One expense as the DETAIL_HEADERS columns. Amount stays numeric
    (float) so xlsx/csv land a real number in the cell — a spreadsheet
    that can't sum its own Amount column is not much of an export."""
    product = expense.product
    return [
        expense.createdAt.strftime("%Y-%m-%d %H:%M"),
        expense.sisterProfile.buyerProfile.name,
        _sister_label(expense.sisterProfile),
        product.name if product else "—",
        product.styleNumber if product else "—",
        expense.get_sourceType_display(),
        expense.fieldName or "—",
        float(expense.amount or 0),
        expense.currency,
        expense.remarks or "",
        expense.createdBy.display_name() if expense.createdBy else "—",
    ]


def totals_by_currency(expenses) -> "OrderedDict[str, Decimal]":
    totals: OrderedDict[str, Decimal] = OrderedDict()
    for expense in expenses:
        totals[expense.currency] = totals.get(expense.currency, Decimal("0")) + (expense.amount or Decimal("0"))
    return OrderedDict(sorted(totals.items()))


def totals_line(expenses) -> str:
    totals = totals_by_currency(expenses)
    if not totals:
        return "0.00"
    return "  +  ".join(f"{_money(amount)} {currency}" for currency, amount in totals.items())


def grouped(expenses, group_by: str) -> "OrderedDict[str, list]":
    """Expenses bucketed by the chosen dimension, groups in alphabetical
    order and rows inside each group oldest-first. `none` yields a single
    unnamed bucket so every renderer can walk the same structure."""
    if group_by == "none" or group_by not in GROUP_BY_OPTIONS:
        return OrderedDict([("", list(expenses))])
    buckets: dict[str, list] = {}
    for expense in expenses:
        buckets.setdefault(group_label(expense, group_by), []).append(expense)
    return OrderedDict(sorted(buckets.items(), key=lambda item: item[0].lower()))


def summary_rows(expenses, group_by: str) -> list[tuple[str, str, int, Decimal]]:
    """(group, currency, row count, total) — one row per group *per
    currency*, never a single blended total per group."""
    rows: list[tuple[str, str, int, Decimal]] = []
    for label, bucket in grouped(expenses, group_by).items():
        per_currency: OrderedDict[str, list] = OrderedDict()
        for expense in bucket:
            per_currency.setdefault(expense.currency, []).append(expense)
        for currency, currency_rows in sorted(per_currency.items()):
            total = sum((e.amount or Decimal("0")) for e in currency_rows)
            rows.append((label, currency, len(currency_rows), total))
    return rows


def expenses_filename(ext: str, *, scope: str = "", date_from: str = "", date_to: str = "") -> str:
    parts = ["expenses"]
    if scope:
        parts.append(slugify(scope)[:40] or "filtered")
    if date_from or date_to:
        parts.append(f"{date_from or 'start'}_to_{date_to or 'today'}")
    return "-".join(p for p in parts if p) + f".{ext}"


# ── CSV ────────────────────────────────────────────────────────────────
# Deliberately plain: header row + data rows and nothing else, so the file
# drops straight into accounting software or a pivot table. The meta/filter
# block that xlsx and pdf carry would break that, and lives in the filename
# instead.

def render_expenses_csv(expenses, group_by: str = "none") -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    grouping = group_by != "none" and group_by in GROUP_BY_OPTIONS
    # Qualified rather than bare, because grouping by e.g. Buyer would
    # otherwise put two identically-named columns in one csv and any pivot
    # table built on it silently picks whichever it saw last.
    group_header = f"Group ({GROUP_BY_OPTIONS[group_by][0]})" if grouping else ""
    writer.writerow(([group_header] if grouping else []) + DETAIL_HEADERS)
    for label, bucket in grouped(expenses, group_by).items():
        for expense in bucket:
            writer.writerow(([label] if grouping else []) + _detail_row(expense))
    # BOM: Excel opens a BOM-less UTF-8 csv as the system codepage and
    # mangles every non-ASCII buyer/product name in the file.
    return buf.getvalue().encode("utf-8-sig")


# ── XLSX ───────────────────────────────────────────────────────────────

def render_expenses_xlsx(expenses, meta: list[tuple[str, str]], group_by: str = "none") -> bytes:
    wb = Workbook()
    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF", size=13)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    line_header_font = Font(bold=True)
    line_header_fill = PatternFill("solid", fgColor=LINE_HEADER_FILL)
    group_fill = PatternFill("solid", fgColor=GROUP_FILL)
    money_format = "#,##0.00"

    # ── Sheet 1: Summary (filters that produced the file + totals) ──
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="EXPENSE REPORT")
    title.font = header_font
    title.fill = header_fill
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Grand Total").font = bold
    ws.cell(row=row, column=2, value=totals_line(expenses))
    row += 1
    ws.cell(row=row, column=1, value="Rows").font = bold
    ws.cell(row=row, column=2, value=len(expenses))
    row += 2

    if group_by != "none" and group_by in GROUP_BY_OPTIONS:
        group_header = GROUP_BY_OPTIONS[group_by][0]
        for col, text in enumerate([group_header, "Currency", "Entries", "Total"], start=1):
            c = ws.cell(row=row, column=col, value=text)
            c.font = line_header_font
            c.fill = line_header_fill
        row += 1
        for label, currency, count, total in summary_rows(expenses, group_by):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=currency)
            ws.cell(row=row, column=3, value=count)
            ws.cell(row=row, column=4, value=float(total)).number_format = money_format
            row += 1

    for i, width in enumerate([28, 42, 12, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 2: Expenses (every row, with subtotals when grouped) ──
    ds = wb.create_sheet("Expenses")
    grouping = group_by != "none" and group_by in GROUP_BY_OPTIONS
    headers = ([GROUP_BY_OPTIONS[group_by][0]] if grouping else []) + DETAIL_HEADERS
    for col, text in enumerate(headers, start=1):
        c = ds.cell(row=1, column=col, value=text)
        c.font = line_header_font
        c.fill = line_header_fill
    amount_col = headers.index("Amount") + 1

    drow = 2
    for label, bucket in grouped(expenses, group_by).items():
        for expense in bucket:
            values = ([label] if grouping else []) + _detail_row(expense)
            for col, value in enumerate(values, start=1):
                cell = ds.cell(row=drow, column=col, value=value)
                if col == amount_col:
                    cell.number_format = money_format
            drow += 1
        if grouping:
            for currency, total in totals_by_currency(bucket).items():
                subtotal = ds.cell(row=drow, column=1, value=f"{label} — subtotal ({currency})")
                subtotal.font = bold
                subtotal.fill = group_fill
                c = ds.cell(row=drow, column=amount_col, value=float(total))
                c.font = bold
                c.fill = group_fill
                c.number_format = money_format
                ds.cell(row=drow, column=amount_col + 1, value=currency).font = bold
                drow += 1
            drow += 1

    ds.freeze_panes = "A2"
    ds.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    widths = ([24] if grouping else []) + [17, 22, 20, 24, 14, 20, 18, 13, 10, 30, 18]
    for i, width in enumerate(widths, start=1):
        ds.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────────────
# Landscape A4: eleven columns of expense detail do not fit portrait
# without shrinking the type past readable.

def render_expenses_pdf(expenses, meta: list[tuple[str, str]], group_by: str = "none") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ExpenseTitle", parent=styles["Title"], fontSize=16, spaceAfter=2)
    section_style = ParagraphStyle("Section", parent=styles["Heading3"], spaceBefore=10, spaceAfter=4)
    # Text columns render through Paragraph so long product names and
    # remarks wrap inside their cell instead of overprinting the next one.
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7.5, leading=9, wordWrap="CJK")
    header_style = ParagraphStyle("CellHeader", parent=cell_style, fontName="Helvetica-Bold")

    story = [Paragraph("Expense Report", title_style), Spacer(1, 6)]

    meta_rows = [[Paragraph(f"<b>{_esc(label)}</b>", cell_style), Paragraph(_esc(value), cell_style)]
                 for label, value in meta]
    meta_rows.append([Paragraph("<b>Grand Total</b>", cell_style), Paragraph(totals_line(expenses), cell_style)])
    meta_rows.append([Paragraph("<b>Rows</b>", cell_style), Paragraph(str(len(expenses)), cell_style)])
    meta_table = Table(meta_rows, colWidths=[40 * mm, 130 * mm])
    meta_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
    ]))
    story.append(meta_table)

    grouping = group_by != "none" and group_by in GROUP_BY_OPTIONS
    if grouping:
        story.append(Paragraph(f"Summary by {GROUP_BY_OPTIONS[group_by][0]}", section_style))
        summary_data = [[Paragraph(h, header_style) for h in (GROUP_BY_OPTIONS[group_by][0], "Currency", "Entries", "Total")]]
        for label, currency, count, total in summary_rows(expenses, group_by):
            summary_data.append([
                Paragraph(_esc(label), cell_style), Paragraph(_esc(currency), cell_style),
                Paragraph(str(count), cell_style), Paragraph(_money(total), cell_style),
            ])
        summary_table = Table(summary_data, colWidths=[90 * mm, 25 * mm, 25 * mm, 35 * mm], repeatRows=1)
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(LINE_HEADER_FILL_HEX)),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(summary_table)

    story.append(Paragraph("Expenses", section_style))
    # Sums to 269mm — the printable width of landscape A4 at 14mm margins.
    col_widths = [24 * mm, 30 * mm, 28 * mm, 34 * mm, 20 * mm, 28 * mm, 24 * mm, 22 * mm, 15 * mm, 24 * mm, 20 * mm]
    detail_data = [[Paragraph(h, header_style) for h in DETAIL_HEADERS]]
    group_row_indexes: list[int] = []
    for label, bucket in grouped(expenses, group_by).items():
        if grouping:
            group_row_indexes.append(len(detail_data))
            detail_data.append([
                Paragraph(f"<b>{_esc(label)}</b> — {totals_line(bucket)} ({len(bucket)} entries)", cell_style),
            ] + [""] * (len(DETAIL_HEADERS) - 1))
        for expense in bucket:
            values = _detail_row(expense)
            detail_data.append([
                Paragraph(_money(v) if isinstance(v, float) else _esc(v), cell_style) for v in values
            ])

    detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
    detail_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(LINE_HEADER_FILL_HEX)),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (7, 1), (7, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]
    for index in group_row_indexes:
        detail_style.append(("BACKGROUND", (0, index), (-1, index), colors.HexColor("#F1F5F9")))
        detail_style.append(("SPAN", (0, index), (-1, index)))
    detail_table.setStyle(TableStyle(detail_style))
    story.append(detail_table)

    doc.build(story)
    return buf.getvalue()
