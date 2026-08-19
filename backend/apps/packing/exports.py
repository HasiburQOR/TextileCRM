"""
Packing List document exports (Excel + PDF) — both carry full buyer
identity (name, branding, contact info), the Sister Profile / PO this list
was built against, every carton row, and the list's totals. Nothing here
is computed — it only formats fields the PackingList/PackingCarton models
already computed via apps.packing.services.

Packing_List_Module_Instructions.md §6: "column headers must be
localizable, not hardcoded English strings in the template." No EN/KA
translation pipeline actually exists anywhere in this codebase yet (no
WeasyPrint, no gettext/django.utils.translation — the doc's assumption that
one already exists doesn't hold here), so real bilingual output is out of
scope for this change. CARTON_ROW_HEADERS below is the minimal honest
version of that ask: one label lookup both renderers share, so adding
actual localization later is a lookup-table edit, not a template rewrite.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.packing.models import PackingCarton, PackingList

HEADER_FILL = "1E2937"
LINE_HEADER_FILL = "E5E7EB"
LINE_HEADER_FILL_HEX = "#" + LINE_HEADER_FILL

# Shared column set for the carton-row table in both exports — order here
# is the order columns render in, in both renderers. Columns merged across
# a Style/Product group (see _group_by_product below) are marked so both
# renderers apply the same merge without duplicating the list of which
# columns count as "style-level".
CARTON_ROW_COLUMNS = [
    ("styleNo", "Style No", True),
    ("product", "Product", True),
    ("poNo", "PO No", True),
    ("colorName", "Color", False),
    ("patternNo", "Pattern No", False),
    ("cartonNoFrom", "CTN From", False),
    ("cartonNoTo", "CTN To", False),
    ("noOfCartons", "CTNS", False),
    ("orderQty", "Order Qty", False),
    ("shipQty", "Ship Qty", False),
    ("sizeBreakdown", "Size Breakdown", False),
    ("totalPcsPerCarton", "PC/CTN", False),
    ("totalGrossWeight", "TTL G.W (kg)", False),
    ("totalNetWeight", "TTL N.W (kg)", False),
    ("totalCbm", "CBM", False),
]
CARTON_ROW_HEADERS = [label for _, label, _ in CARTON_ROW_COLUMNS]
# 1-based column indices (openpyxl) / 0-based (reportlab, computed at use).
_MERGED_COLUMN_INDICES_1BASED = [i + 1 for i, (_, _, merged) in enumerate(CARTON_ROW_COLUMNS) if merged]


def _num(value, decimals=2) -> str:
    return f"{float(value or 0):,.{decimals}f}"


def _size_breakdown_text(size_breakdown: list) -> str:
    """Custom_Size_Breakdown_Feature.md: size_breakdown is a free-form array
    of {size_label, quantity}, not a fixed dict."""
    return " / ".join(f"{e.get('size_label', '')}:{e.get('quantity', 0)}" for e in (size_breakdown or [])) or "—"


def _carton_row_values(c: PackingCarton) -> list:
    return [
        c.styleNo or c.product.styleNumber,
        c.product.name,
        c.poNo or "—",
        c.colorName or "—",
        c.patternNo or "—",
        c.cartonNoFrom,
        c.cartonNoTo,
        c.noOfCartons,
        c.orderQty,
        c.shipQty,
        _size_breakdown_text(c.sizeBreakdown),
        c.totalPcsPerCarton,
        float(c.totalGrossWeight),
        float(c.totalNetWeight),
        float(c.totalCbm),
    ]


def _group_by_product(cartons: list[PackingCarton]) -> list[tuple[int, int]]:
    """Packing_List_Module_Instructions.md §6: 'preserve the visual
    grouping (Style No and PO No shown once per group, not repeated on
    every row)'. Returns (start_index, end_index) row-index pairs (into
    `cartons`, 0-based, inclusive) for each run of *consecutive* rows
    sharing the same Product — never a global re-sort/re-group, since
    carton ranges are already sequential per color/style in normal use."""
    groups = []
    start = 0
    for i in range(1, len(cartons) + 1):
        if i == len(cartons) or cartons[i].product_id != cartons[start].product_id:
            groups.append((start, i - 1))
            start = i
    return groups


def packing_list_filename(pl: PackingList, ext: str) -> str:
    return f"{pl.poNo or str(pl.id)}.{ext}"


def _buyer_block(pl: PackingList) -> list[tuple[str, str]]:
    sister_profile = pl.sisterProfile
    buyer = sister_profile.buyerProfile
    return [
        ("Buyer", buyer.name),
        ("Branding", buyer.branding or "—"),
        ("Contact Info", (buyer.contactInfo or "—").replace("\n", ", ")),
        ("Sister Profile / PO", sister_profile.poReference or str(sister_profile.id)),
    ]


def _list_meta(pl: PackingList) -> list[tuple[str, str]]:
    rows = [
        ("PO No", pl.poNo or "—"),
        ("Brand Name", pl.brandName or "—"),
        ("Date", pl.date.strftime("%Y-%m-%d") if pl.date else "—"),
    ]
    if pl.frontMark:
        rows.append(("Front Mark", pl.frontMark))
    if pl.sideMark:
        rows.append(("Side Mark", pl.sideMark))
    return rows


def render_packing_list_xlsx(pl: PackingList) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"

    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF", size=13)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    line_header_font = Font(bold=True)
    line_header_fill = PatternFill("solid", fgColor=LINE_HEADER_FILL)
    group_alignment = Alignment(vertical="center")

    row = 1
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value=f"PACKING LIST — {pl.poNo or pl.id}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26
    row += 2

    for label, value in _buyer_block(pl) + _list_meta(pl):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    for col, text in enumerate(CARTON_ROW_HEADERS, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = line_header_font
        c.fill = line_header_fill
    table_start_row = row + 1
    row += 1

    cartons = list(pl.cartons.all())
    for c in cartons:
        for col, value in enumerate(_carton_row_values(c), start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Merge Style No / Product / PO No cells down each consecutive
    # same-Product run — never flattened into repeated values per row.
    for start_idx, end_idx in _group_by_product(cartons):
        if end_idx == start_idx:
            continue
        first_row = table_start_row + start_idx
        last_row = table_start_row + end_idx
        for col in _MERGED_COLUMN_INDICES_1BASED:
            ws.merge_cells(start_row=first_row, start_column=col, end_row=last_row, end_column=col)
            ws.cell(row=first_row, column=col).alignment = group_alignment

    row += 1

    totals = [
        ("Total Order Qty", pl.totalOrderQty), ("Total Ship Qty", pl.totalShipQty),
        ("Short/Excess Qty", pl.shortExcessQty), ("Short/Excess %", f"{pl.shortExcessPct}%"),
        ("Total Carton Qty", pl.totalCartonQty), ("Total Gross Weight (kg)", _num(pl.totalGrossWeight)),
        ("Total Net Weight (kg)", _num(pl.totalNetWeight)), ("Total CBM", _num(pl.totalCbm, 4)),
    ]
    for label, value in totals:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    widths = [14, 24, 14, 22, 12, 10, 8, 8, 10, 10, 22, 8, 13, 13, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_packing_list_pdf(pl: PackingList) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PLTitle", parent=styles["Title"], fontSize=17, spaceAfter=2)
    section_style = ParagraphStyle("Section", parent=styles["Heading3"], spaceBefore=8, spaceAfter=4)

    story = [
        Paragraph(f"Packing List — {pl.poNo or pl.id}", title_style),
        Spacer(1, 6),
    ]

    info_rows = [[Paragraph(f"<b>{l}</b>", styles["Normal"]), Paragraph(str(v), styles["Normal"])]
                 for l, v in (_buyer_block(pl) + _list_meta(pl))]
    info_table = Table(info_rows, colWidths=[40 * mm, 110 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
    ]))
    story.append(info_table)

    story.append(Paragraph("Carton Rows", section_style))
    cartons = list(pl.cartons.all())
    data = [CARTON_ROW_HEADERS]
    for c in cartons:
        row_values = _carton_row_values(c)
        # Format the decimal columns for display (raw model values are what
        # _carton_row_values returns, shared with the XLSX renderer which
        # wants numbers rather than pre-formatted strings).
        row_values[12] = _num(row_values[12])
        row_values[13] = _num(row_values[13])
        row_values[14] = _num(row_values[14], 4)
        data.append([str(v) for v in row_values])

    col_widths = [16, 22, 14, 20, 14, 10, 10, 8, 10, 10, 20, 8, 13, 13, 10]
    col_widths = [w * mm for w in col_widths]
    table_style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(LINE_HEADER_FILL_HEX)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.2),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    # Same merge-consecutive-same-Product-runs rule as the XLSX renderer —
    # reportlab's SPAN is 0-based (col, row), row 0 is the header.
    for start_idx, end_idx in _group_by_product(cartons):
        if end_idx == start_idx:
            continue
        first_row, last_row = start_idx + 1, end_idx + 1
        for col in (0, 1, 2):  # Style No, Product, PO No
            table_style_commands.append(("SPAN", (col, first_row), (col, last_row)))
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(table_style_commands))
    story.append(table)

    story.append(Paragraph("Totals", section_style))
    totals_rows = [
        ["Total Order Qty", str(pl.totalOrderQty), "Total Ship Qty", str(pl.totalShipQty)],
        ["Short/Excess Qty", f"{pl.shortExcessQty} ({pl.shortExcessPct}%)", "Total Cartons", str(pl.totalCartonQty)],
        ["Total Gross Weight", f"{_num(pl.totalGrossWeight)} kg", "Total Net Weight", f"{_num(pl.totalNetWeight)} kg"],
        ["Total CBM", _num(pl.totalCbm, 4), "", ""],
    ]
    totals_table = Table(totals_rows, colWidths=[38 * mm, 38 * mm, 38 * mm, 38 * mm])
    totals_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(totals_table)

    doc.build(story)
    return buf.getvalue()
